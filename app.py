# เหลือทำ cancel
# การคำนวณ discout
# เลือกเป็นช่วงผู้โดยสาร
# เลือก bike ได้หลายคัน

from flask import Flask, render_template, request, jsonify, url_for, redirect, send_file, session, flash
import os
import re
import io
from datetime import date, datetime, time
import psycopg2
from psycopg2.extras import RealDictCursor
import uuid
import openpyxl
import tempfile
import subprocess
import platform
from openpyxl.drawing.image import Image

from functools import wraps
import secrets

app = Flask(__name__)

#---------------------------------------------- LOGIN ------------------------------------------------------
# สร้าง secret key สำหรับ session
app.secret_key = 'booking_system_secret_key_2024'

# Decorator สำหรับตรวจสอบการ login
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            flash('Please log in to continue', 'error')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

# ---------------------------------------Connect database ---------------------------------------

# Database connection
# def get_db_connection():
#     database_url = os.getenv("DATABASE_URL")
    
#     if database_url:
#         # สำหรับ Render.com
#         return psycopg2.connect(database_url)
#     else:
#         # สำหรับ local development
#         return psycopg2.connect(
#             host=os.getenv("DB_HOST"),
#             port=int(os.getenv("DB_PORT", 5432)),
#             user=os.getenv("DB_USER"),
#             password=os.getenv("DB_PASS"),
#             database=os.getenv("DB_NAME")
#         )

def get_db_connection():
    return psycopg2.connect(
        host="dpg-d0qsdf95pdvs73atfls0-a.oregon-postgres.render.com",  # ✅ host ต้องไม่มี protocol
        port="5432",
        dbname="booking_system_mmdr",
        user="booking_user",
        password="1YtEzFr8UkRTNzzwYtKQe8jaaremuxyA",
        sslmode="require"  # ✅ ใช้ sslmode=required กับ Render
    )


def load_data_from_file(filename):
    try:
        script_dir = os.path.dirname(os.path.abspath(__file__))
        file_path = os.path.join(script_dir, filename)
        
        with open(file_path, 'r') as f:
            data = [line.strip() for line in f.readlines() if line.strip()]
        
        return data
    except Exception as e:
        print(f"Error loading file {filename}: {str(e)}")
        return []


@app.route("/")
def index():
    # ถ้า login แล้วให้ไปหน้า home
    if 'user_id' in session:
        return redirect(url_for('home'))
    return render_template("P00_login.html")



@app.route("/process_login", methods=["POST"])
def process_login():
    try:
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        
        if not username or not password:
            return jsonify({
                "success": False, 
                "message": "Please enter both username and password."
            })
        
        conn = get_db_connection()
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        
        # Query เพื่อดึงข้อมูล user รวมทั้ง role
        query = "SELECT id, username, pass, first_name, last_name, role FROM login WHERE username = %s"
        cursor.execute(query, (username,))
        user = cursor.fetchone()
        
        cursor.close()
        conn.close()
        
        # เปรียบเทียบ password แบบ plain text
        if user and user['pass'] == password:
            # เก็บข้อมูลใน session รวมทั้ง role
            session['user_id'] = user['id']
            session['username'] = user['username']
            session['first_name'] = user['first_name']
            session['last_name'] = user['last_name']
            session['full_name'] = f"{user['first_name']} {user['last_name']}"
            session['role'] = user.get('role', 'staff')  # เพิ่ม role, default เป็น staff
            
            return jsonify({
                "success": True, 
                "message": "Login successful.",
                "redirect_url": url_for('home')
            })
        else:
            return jsonify({
                "success": False, 
                "message": "Invalid username or password."
            })
            
    except Exception as e:
        return jsonify({
            "success": False, 
            "message": f"An error occurred: {str(e)}"
        })

# เพิ่ม route สำหรับ logout
@app.route("/logout")
def logout():
    session.clear()
    flash('Logout Successful', 'success')
    return redirect(url_for('index'))

@app.route("/home")
@login_required
def home():
    return render_template("P0_home.html")

# -------------------------------- Tour Page (Updated) --------------------------------------------------------------

@app.route("/tour_rental")
@login_required
def tour_rental():
    # เชื่อมต่อกับฐานข้อมูล
    conn = get_db_connection()
    cursor = conn.cursor()
    
    # ดึงข้อมูลห้องจากตาราง room
    query = "SELECT room FROM room ORDER BY room"
    cursor.execute(query)
    room_results = cursor.fetchall()
    rooms = [room[0] for room in room_results]
    
    # ดึงข้อมูล tour companies เท่านั้น (ไม่ต้อง motorbike)
    query_tour = "SELECT DISTINCT company_name FROM tour_tabel ORDER BY company_name"
    cursor.execute(query_tour)
    tour_companies = [company[0] for company in cursor.fetchall()]
    
    cursor.close()
    conn.close()
    
    today = date.today().isoformat()  # YYYY-MM-DD
    
    return render_template('P1_tour_rental.html', 
                         room=rooms, 
                         today=today,
                         tour_companies=tour_companies)

# เพิ่ม API endpoint สำหรับดึงข้อมูล details ตาม company และ experience type
@app.route("/get_company_details", methods=["POST"])
@login_required
def get_company_details():
    try:
        experience_type = request.form.get('experience_type')
        company_name = request.form.get('company_name')
        
        if not experience_type or not company_name:
            return jsonify({"success": False, "message": "Missing required parameters"})
        
        conn = get_db_connection()
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        
        # เลือกตารางตาม experience type
        if experience_type == 'tour':
            table_name = 'tour_tabel'
        elif experience_type == 'motorbike':
            table_name = 'motorbike_tabel'
        else:
            return jsonify({"success": False, "message": "Invalid experience type"})
        
        # ดึงข้อมูล details สำหรับ company ที่เลือก
        query = f"SELECT detail, received FROM {table_name} WHERE company_name = %s ORDER BY detail"
        cursor.execute(query, (company_name,))
        details = cursor.fetchall()
        
        cursor.close()
        conn.close()
        
        return jsonify({"success": True, "details": details})
        
    except Exception as e:
        return jsonify({"success": False, "message": f"Error: {str(e)}"})

# เพิ่ม API endpoint สำหรับดึงราคาตาม detail ที่เลือก
@app.route("/get_detail_price", methods=["POST"])
@login_required
def get_detail_price():
    try:
        experience_type = request.form.get('experience_type')
        detail = request.form.get('detail')
        
        if not experience_type or not detail:
            return jsonify({"success": False, "message": "Missing required parameters"})
        
        conn = get_db_connection()
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        
        # เลือกตารางตาม experience type
        if experience_type == 'tour':
            table_name = 'tour_tabel'
        elif experience_type == 'motorbike':
            table_name = 'motorbike_tabel'
        else:
            return jsonify({"success": False, "message": "Invalid experience type"})
        
        # ดึงข้อมูลราคาสำหรับ detail ที่เลือก
        query = f"SELECT received FROM {table_name} WHERE detail = %s"
        cursor.execute(query, (detail,))
        result = cursor.fetchone()
        
        cursor.close()
        conn.close()
        
        if result:
            return jsonify({"success": True, "price": result['received']})
        else:
            return jsonify({"success": False, "message": "Detail not found"})
        
    except Exception as e:
        return jsonify({"success": False, "message": f"Error: {str(e)}"})
    
# เพิ่ม API endpoint สำหรับดึงข้อมูล companies
@app.route("/api/companies", methods=["GET"])
@login_required
def get_companies():
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # ดึงข้อมูล tour companies
        query_tour = "SELECT DISTINCT company_name FROM tour_tabel ORDER BY company_name"
        cursor.execute(query_tour)
        tour_companies = [company[0] for company in cursor.fetchall()]
        
        # ดึงข้อมูล motorbike companies
        query_motorbike = "SELECT DISTINCT company_name FROM motorbike_tabel ORDER BY company_name"
        cursor.execute(query_motorbike)
        motorbike_companies = [company[0] for company in cursor.fetchall()]
        
        cursor.close()
        conn.close()
        
        return jsonify({
            "success": True,
            "tour_companies": tour_companies,
            "motorbike_companies": motorbike_companies
        })
        
    except Exception as e:
        return jsonify({
            "success": False,
            "message": f"Error: {str(e)}",
            "tour_companies": [],
            "motorbike_companies": []
        })
    
@app.context_processor
def inject_user_info():
    """ส่งข้อมูล user ปัจจุบันไปยัง template ทุกหน้า"""
    user_info = {
        'user_id': session.get('user_id'),
        'username': session.get('username'),
        'first_name': session.get('first_name'),
        'last_name': session.get('last_name'),
        'full_name': session.get('full_name'),
        'role': session.get('role')
    }
    return dict(current_user=user_info)

@app.route("/api/current_user", methods=["GET"])
@login_required
def get_current_user():
    """API สำหรับดึงข้อมูล user ปัจจุบัน"""
    try:
        return jsonify({
            "success": True,
            "user": {
                "user_id": session.get('user_id'),
                "username": session.get('username'),
                "first_name": session.get('first_name'),
                "last_name": session.get('last_name'),
                "full_name": session.get('full_name'),
                "role": session.get('role')
            }
        })
    except Exception as e:
        return jsonify({
            "success": False,
            "message": f"Error: {str(e)}"
        })
    

@app.route("/submit_tour_booking", methods=["POST"])
@login_required
def submit_tour_booking():
    try:
        # Connect to the database to get the latest booking number
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Get current year and month (in format YY and MM)
        current_date = datetime.now()
        year_prefix = current_date.strftime("%y")  # 2-digit year (e.g., 25 for 2025)
        month_prefix = current_date.strftime("%m")  # 2-digit month (e.g., 05 for May)
        prefix = f"{year_prefix}{month_prefix}"  # e.g., 2505
        
        # Query to find the highest booking number with the current year-month prefix
        # ⚠️ เปลี่ยนจาก tour_motobike_rental เป็น tour_rental
        query = """
        SELECT booking_no FROM tour_rental 
        WHERE booking_no LIKE %s 
        ORDER BY CAST(SUBSTRING(booking_no, 5) AS INTEGER) DESC 
        LIMIT 1
        """
        cursor.execute(query, (f"{prefix}%",))
        result = cursor.fetchone()
        
        if result:
            last_booking_no = result[0]
            try:
                running_number = int(last_booking_no[4:]) + 1
            except (ValueError, IndexError):
                running_number = 1
        else:
            running_number = 1
        
        # สร้าง booking number และตรวจสอบว่าไม่ซ้ำ
        max_attempts = 1000
        booking_no = None
        
        for attempt in range(max_attempts):
            temp_booking_no = f"T{prefix}{running_number:05d}"
            
            # ⚠️ เปลี่ยนจาก tour_motobike_rental เป็น tour_rental
            check_query = "SELECT booking_no FROM tour_rental WHERE booking_no = %s"
            cursor.execute(check_query, (temp_booking_no,))
            exists = cursor.fetchone()
            
            if not exists:
                booking_no = temp_booking_no
                break
            else:
                running_number += 1
        
        if not booking_no:
            return jsonify({"success": False, "message": "Unable to generate unique booking number"})

        # Get form data (เหมือนเดิม)
        customer_name = request.form.get('name', '')
        customer_surname = request.form.get('surname', '')
        room = request.form.get('room', '')
        # ⚠️ ไม่ต้องดึง experience_type แล้ว
        company_name = request.form.get('company', '')
        detail = request.form.get('detail', '')
        pickup_time = request.form.get('time', '')
        travel_date = request.form.get('date', '')
        quantity = request.form.get('persons', '0')
        price_per_person = request.form.get('price', '0')
        payment_status = request.form.get('status', '')
        staff_name = request.form.get('staffName', '')
        payment_method = request.form.get('paymentmethod', '') # อย่าลืมใส่ variable = paymentmethod ตัวนี้ใน tour_rental และ tour_transfer
        remark = request.form.get('remark', '') # อย่าลืมใส่ variable = remark ตัวนี้ใน tour_rental และ tour_transfer
        discount = request.form.get('discount', '')
        
        # Calculate total receivedtotal
        received = float(request.form.get('total', ''))
        # received = float(price_per_person) * int(quantity) if price_per_person and quantity else 0
        booking_date = date.today().isoformat()
        
        # Validate required fields (ลบ experience_type ออก)
        if not all([customer_name, customer_surname, room, 
                   company_name, detail, pickup_time, travel_date, 
                   quantity, price_per_person, payment_status, staff_name, payment_method]):
            cursor.close()
            conn.close()
            return jsonify({"success": False, "message": "Please fill in all required fields"})
        
        # ⚠️ Insert ลงตาราง tour_rental
        query = """
        INSERT INTO tour_rental (
            travel_date, pickup_time, booking_date, booking_no, 
            customer_name, customer_surname, room, company_name, 
            detail, quantity, received, payment_status, 
            staff_name, payment_method, remark, discount, price
        ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """
        
        values = (
            travel_date, pickup_time, booking_date, booking_no,
            customer_name, customer_surname, room, company_name,
            detail, quantity, received, payment_status,
            staff_name, payment_method, remark, discount, price_per_person
        )
        
        cursor.execute(query, values)
        conn.commit()
        
        cursor.close()
        conn.close()
        
        return jsonify({
            "success": True, 
            "message": "Tour booking successfully submitted", 
            "booking_no": booking_no,
            "total_amount": received
        })
    
    except Exception as e:
        try:
            if 'cursor' in locals():
                cursor.close()
            if 'conn' in locals():
                conn.close()
        except:
            pass
        
        return jsonify({"success": False, "message": f"Error: {str(e)}"})

@app.route("/tour_rental_form")
@login_required
def view_tour_bookings():
    try:
        conn = get_db_connection()
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        
        # ดึงข้อมูล bookings
        query = """
        SELECT * FROM tour_rental
        WHERE travel_date >= CURRENT_DATE
        ORDER BY travel_date ASC
        LIMIT 50
        """
        cursor.execute(query)
        bookings = cursor.fetchall()
        
        # ⚠️ เพิ่มส่วนนี้ - ดึง booking_list สำหรับ dropdown
        cursor.execute("SELECT DISTINCT booking_no, customer_name, customer_surname FROM tour_rental ORDER BY booking_no DESC")
        booking_list = cursor.fetchall()
        
        # แปลงวันที่และเวลา (โค้ดเดิม...)
        for booking in bookings:
            if booking['travel_date']:
                booking['travel_date'] = booking['travel_date'].strftime('%d/%m/%Y')
            if booking['booking_date']:
                booking['booking_date'] = booking['booking_date'].strftime('%d/%m/%Y')
            
            # แก้ไขส่วนนี้ - ตรวจสอบว่า pickup_time เป็น datetime.time ไม่ใช่ timedelta
            if booking['pickup_time']:
                if hasattr(booking['pickup_time'], 'strftime'):
                    booking['pickup_time'] = booking['pickup_time'].strftime('%H:%M')
                else:
                    # ถ้าเป็น timedelta ให้แปลงเป็นรูปแบบชั่วโมง:นาที
                    total_seconds = int(booking['pickup_time'].total_seconds())
                    hours = total_seconds // 3600
                    minutes = (total_seconds % 3600) // 60
                    booking['pickup_time'] = f"{hours:02d}:{minutes:02d}"
        
        cursor.close()
        conn.close()
        
        # ⚠️ เพิ่ม booking_list ในการ return
        return render_template('P1_tour_form.html', 
                             bookings=bookings,
                             booking_list=booking_list) 
    
    except Exception as e:
        return f"Error: {str(e)}"
    
# ---- อัพเดต Context Processor ----
@app.context_processor
def inject_room_list():
    # เชื่อมต่อกับฐานข้อมูล
    conn = get_db_connection()
    cursor = conn.cursor()
    
    # ดึงข้อมูลห้องจากตาราง room
    query = "SELECT room FROM room ORDER BY room"
    cursor.execute(query)
    
    # ดึงผลลัพธ์ออกมาเป็นรายการธรรมดา
    room_results = cursor.fetchall()
    room_list = [room[0] for room in room_results]
    
    # ดึงข้อมูล companies สำหรับใช้ใน edit modal
    query_tour = "SELECT DISTINCT company_name FROM tour_tabel ORDER BY company_name"
    cursor.execute(query_tour)
    tour_companies = [company[0] for company in cursor.fetchall()]
    
    query_motorbike = "SELECT DISTINCT company_name FROM motorbike_tabel ORDER BY company_name"
    cursor.execute(query_motorbike)
    motorbike_companies = [company[0] for company in cursor.fetchall()]
    
    cursor.close()
    conn.close()
    
    return dict(
        room_list=room_list,
        tour_companies=tour_companies,
        motorbike_companies=motorbike_companies
    )
        

# ---- อัพเดต Search Bookings ----
@app.route("/search_tour_bookings", methods=["GET", "POST"])
@login_required
def search_tour_bookings():
    try:
        # ดึง booking list และ room list สำหรับ dropdown (ใช้ทั้ง GET และ POST)
        conn = get_db_connection()
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        
        # ดึง booking list สำหรับ dropdown
        cursor.execute("SELECT DISTINCT booking_no, customer_name, customer_surname FROM tour_rental ORDER BY booking_no DESC")
        booking_list = cursor.fetchall()
        
        cursor.close()
        conn.close()
        
        if request.method == "POST":
            # รับค่าจากฟอร์มค้นหา
            start_date = request.form.get('start_date', '')
            end_date = request.form.get('end_date', '')
            booking_no = request.form.get('booking_no', '')
            name_surname = request.form.get('name_surname', '')
            
            # เชื่อมต่อกับฐานข้อมูล
            conn = get_db_connection()
            cursor = conn.cursor(cursor_factory=RealDictCursor)
            
            # ⚠️ เปลี่ยนจาก tour_motobike_rental เป็น tour_rental
            query = """
            SELECT 
                booking_no, 
                travel_date, 
                pickup_time,
                booking_date,
                customer_name, 
                customer_surname, 
                room, 
                company_name, 
                detail, 
                quantity, 
                received, 
                payment_status,
                payment_method, 
                staff_name,
                remark,
                discount
            FROM tour_rental
            WHERE 1=1
            """
            
            params = []
            
            # เพิ่มเงื่อนไขการค้นหา
            if start_date:
                query += " AND travel_date >= %s"
                params.append(start_date)
            
            if end_date:
                query += " AND travel_date <= %s"
                params.append(end_date)
            
            if booking_no:
                query += " AND booking_no = %s"
                params.append(booking_no)
                
            if name_surname:
                # แยกชื่อและนามสกุล
                name_parts = name_surname.strip().split(' ', 1)
                if len(name_parts) == 2:
                    first_name, last_name = name_parts
                    query += " AND customer_name = %s AND customer_surname = %s"
                    params.extend([first_name, last_name])
                else:
                    # ถ้าไม่สามารถแยกได้ ให้ค้นหาจากชื่อเต็ม
                    query += " AND CONCAT(customer_name, ' ', customer_surname) = %s"
                    params.append(name_surname)
            
            query += " ORDER BY travel_date DESC, booking_no DESC"
            
            cursor.execute(query, params)
            bookings = cursor.fetchall()
            
            # แปลงรูปแบบวันที่และเวลา
            for booking in bookings:
                if booking['travel_date']:
                    booking['travel_date'] = booking['travel_date'].strftime('%d/%m/%Y')
                if booking['booking_date']:
                    booking['booking_date'] = booking['booking_date'].strftime('%d/%m/%Y')
                
                if booking['pickup_time']:
                    if hasattr(booking['pickup_time'], 'strftime'):
                        booking['pickup_time'] = booking['pickup_time'].strftime('%H:%M')
                    else:
                        total_seconds = int(booking['pickup_time'].total_seconds())
                        hours = total_seconds // 3600
                        minutes = (total_seconds % 3600) // 60
                        booking['pickup_time'] = f"{hours:02d}:{minutes:02d}"
            
            cursor.close()
            conn.close()
            
            return render_template('P1_tour_form.html', 
                                 bookings=bookings,
                                 booking_list=booking_list,
                                 search_start_date=start_date, 
                                 search_end_date=end_date, 
                                 search_booking_no=booking_no,
                                 search_name_surname=name_surname)
        else:
            # ⚠️ GET request - แสดงหน้า search form พร้อม dropdown data
            return render_template('view_tour_bookings')
    
    except Exception as e:
        return f"Error: {str(e)}"

@app.route("/cancel_bookings", methods=["POST"])
@login_required
def cancel_bookings():
    try:
        # รับค่า booking_no ที่ถูกเลือกจากฟอร์ม
        selected_bookings = request.form.getlist('selected_bookings')
        
        # รับประเภทจาก form
        booking_type = request.form.get('booking_type', 'tour')  # default เป็น tour
        
        # รับชื่อผู้ cancel
        cancelled_by = request.form.get('cancelled_by', '').strip()
        
        if not selected_bookings:
            return jsonify({"success": False, "message": "No bookings selected"})
        
        if not cancelled_by:
            return jsonify({"success": False, "message": "Cancellation name is required"})
        
        # เชื่อมต่อกับฐานข้อมูล
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # เลือกตารางตามประเภท
        table_name = 'tour_rental' if booking_type == 'tour' else 'motorbike_rental'
        
        # อัพเดต payment_status แทนการลบ
        placeholders = ', '.join(['%s'] * len(selected_bookings))
        query = f"UPDATE {table_name} SET payment_status = %s WHERE booking_no IN ({placeholders})"
        
        # สร้าง status ใหม่
        cancel_status = f"Cancelled by {cancelled_by}"
        params = [cancel_status] + selected_bookings
        
        cursor.execute(query, params)
        conn.commit()
        
        updated_count = cursor.rowcount
        
        cursor.close()
        conn.close()
        
        if updated_count > 0:
            return jsonify({"success": True, "message": f"Successfully cancelled {updated_count} booking(s) by {cancelled_by}"})
        else:
            return jsonify({"success": False, "message": "No bookings were cancelled"})
        
    except Exception as e:
        return jsonify({"success": False, "message": f"Error: {str(e)}"})


# ---- อัพเดต Get Booking Details สำหรับ Edit ----
@app.route("/get_booking_details", methods=["POST"])
@login_required
def get_booking_details():
    try:
        # รับค่า booking_no จากการ request
        booking_no = request.form.get('booking_no')
        
        # รับประเภทจาก form
        booking_type = request.form.get('booking_type', 'tour')  # default เป็น tour
        
        if not booking_no:
            return jsonify({"success": False, "message": "No booking selected"})
        
        # เชื่อมต่อกับฐานข้อมูล
        conn = get_db_connection()
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        
        # ✅ เลือกตารางและ query ตามประเภท
        if booking_type == 'tour':
            main_table = 'tour_rental'
            price_table = 'tour_tabel'
            
            # ดึงข้อมูลการจองพร้อม JOIN เพื่อดึงราคาต่อคนจากตาราง tour_tabel
            query = f"""
            SELECT 
                mr.*,
                pt.received as price_per_person
            FROM {main_table} mr
            LEFT JOIN {price_table} pt ON mr.detail = pt.detail AND mr.company_name = pt.company_name
            WHERE mr.booking_no = %s
            """
            
        else:  # motorbike
            main_table = 'motorbike_rental'
            
            # ✅ สำหรับ motorbike ไม่ต้อง JOIN เพราะข้อมูล price เก็บไว้ใน motorbike_rental แล้ว
            query = f"""
            SELECT 
                *,
                CONCAT(customer_name, ' ', customer_surname) AS full_name
            FROM {main_table}
            WHERE booking_no = %s
            """
        
        cursor.execute(query, (booking_no,))
        booking = cursor.fetchone()

        if booking and booking_type == 'motorbike':
            # ✅ แปลงข้อมูล comma-separated กลับเป็น arrays สำหรับแสดงผล
            try:
                if booking.get('company_name'):
                    booking['company_list'] = [c.strip() for c in booking['company_name'].split(',')]
                if booking.get('detail'):
                    booking['detail_list'] = [d.strip() for d in booking['detail'].split(',')]
                if booking.get('price'):
                    # ✅ แปลง price string เป็น list
                    booking['price_list'] = [float(p.strip()) for p in str(booking['price']).split(',')]
                if booking.get('quantity'):
                    # ✅ แปลง quantity (รองรับทั้ง "1,2,3" และ "{1,2,3}")
                    qty_str = str(booking['quantity']).strip()
                    if qty_str.startswith('{') and qty_str.endswith('}'):
                        qty_str = qty_str[1:-1]  # เอา { } ออก
                    booking['quantity_list'] = [int(q.strip()) for q in qty_str.split(',')]
            except Exception as e:
                print(f"Error parsing motorbike booking details: {str(e)}")
                # ถ้าแปลงไม่ได้ ให้ใช้ข้อมูลเดิม
                pass
        
        cursor.close()
        conn.close()
        
        if not booking:
            return jsonify({"success": False, "message": "Booking not found"})
        
        # แปลงวันที่และเวลาให้อยู่ในรูปแบบที่เหมาะสม (เหมือนเดิม)
        if booking.get('travel_date'):
            booking['travel_date'] = booking['travel_date'].isoformat()
        if booking.get('booking_date'):
            booking['booking_date'] = booking['booking_date'].isoformat()
        
        # แปลงเวลา (เหมือนเดิม)
        if booking.get('pickup_time'):
            if hasattr(booking['pickup_time'], 'strftime'):
                booking['pickup_time'] = booking['pickup_time'].strftime('%H:%M')
            else:
                # ถ้าเป็น timedelta
                total_seconds = int(booking['pickup_time'].total_seconds())
                hours = total_seconds // 3600
                minutes = (total_seconds % 3600) // 60
                booking['pickup_time'] = f"{hours:02d}:{minutes:02d}"
        
        # เพิ่มราคาต่อคนที่คำนวณแล้ว (สำหรับ tour)
        if booking_type == 'tour' and booking.get('price_per_person'):
            booking['price_per_person'] = float(booking['price_per_person'])
        
        # เพิ่ม booking_type กลับไปด้วย
        booking['booking_type'] = booking_type
        
        return jsonify({"success": True, "booking": booking})
        
    except Exception as e:
        return jsonify({"success": False, "message": f"Error: {str(e)}"})
    
# ---- อัพเดต Update Booking ----
@app.route("/update_booking", methods=["POST"])
@login_required
def update_booking():
    try:
        # รับค่าจากฟอร์มแก้ไข
        booking_no = request.form.get('booking_no')
        
        # ⚠️ เพิ่มตัวนี้ - รับประเภทจาก form
        booking_type = request.form.get('booking_type', 'tour')
        
        customer_name = request.form.get('name', '')
        customer_surname = request.form.get('surname', '')
        room = request.form.get('room', '')
        company_name = request.form.get('company', '')
        detail = request.form.get('detail', '')
        pickup_time = request.form.get('time', '')
        travel_date = request.form.get('date', '')
        quantity = request.form.get('persons', '0')
        price_per_person = request.form.get('price', '0')
        payment_status = request.form.get('status', '')
        payment_method = request.form.get('paymentmethod', '') # อย่าลืมใส่ variable = paymentmethod ตัวนี้ใน tour_rental และ tour_transfer
        remark = request.form.get('remark', '') # อย่าลืมใส่ variable = remark ตัวนี้ใน tour_rental และ tour_transfer
        staff_name = request.form.get('staffName', '')
        discount = request.form.get('discount', '')
        
        if not booking_no:
            return jsonify({"success": False, "message": "Booking number is required"})
        
        # คำนวณ received ใหม่
        received = float(request.form.get('total', ''))
        # received = float(price_per_person) * int(quantity) if price_per_person and quantity else 0
        
        # เชื่อมต่อกับฐานข้อมูล
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # ⚠️ เลือกตารางตามประเภท
        table_name = 'tour_rental' if booking_type == 'tour' else 'motorbike_rental'
        
        # อัปเดตข้อมูลในฐานข้อมูล
        query = f"""
        UPDATE {table_name} SET
            travel_date = %s,
            pickup_time = %s,
            customer_name = %s,
            customer_surname = %s,
            room = %s,
            company_name = %s,
            detail = %s,
            quantity = %s,
            received = %s,
            payment_status = %s,
            payment_method = %s,
            staff_name = %s,
            remark = %s,
            discount = %s,
            price = %s
        WHERE booking_no = %s
        """
        
        values = (
            travel_date, pickup_time, customer_name, customer_surname,
            room, company_name, detail, quantity, received,
            payment_status, payment_method, staff_name, remark, discount, price_per_person, booking_no
        )
        
        cursor.execute(query, values)
        conn.commit()
        
        updated = cursor.rowcount > 0
        
        cursor.close()
        conn.close()
        
        if updated:
            return jsonify({"success": True, "message": "Booking successfully updated"})
        else:
            return jsonify({"success": False, "message": "No changes were made or booking not found"})
        
    except Exception as e:
        return jsonify({"success": False, "message": f"Error: {str(e)}"})
    
# ---- EXCEL FORM GENERATION FUNCTIONALITY ----
def convert_excel_to_pdf_libreoffice(excel_path):
    """Convert Excel file to PDF using LibreOffice headless"""
    
    try:
        # สร้าง PDF path
        pdf_path = excel_path.replace(".xlsx", ".pdf")
        
        # แปลงเป็น absolute paths
        excel_path = os.path.abspath(excel_path)
        pdf_path = os.path.abspath(pdf_path)
        
        # ดึง directory ที่จะส่งออก PDF
        output_dir = os.path.dirname(pdf_path)
        
        print(f"Converting: {excel_path}")
        print(f"Output dir: {output_dir}")
        print(f"Expected PDF: {pdf_path}")
        
        # รัน LibreOffice headless command
        command = [
            'libreoffice',
            '--headless',
            '--convert-to', 'pdf',
            '--outdir', output_dir,
            excel_path
        ]
        
        # รันคำสั่งและรอให้เสร็จ
        result = subprocess.run(
            command,
            capture_output=True,
            text=True,
            timeout=60  # timeout 60 วินาที
        )
        
        print(f"LibreOffice command: {' '.join(command)}")
        print(f"Return code: {result.returncode}")
        print(f"STDOUT: {result.stdout}")
        print(f"STDERR: {result.stderr}")
        
        # ตรวจสอบว่าคำสั่งสำเร็จ
        if result.returncode == 0:
            # ตรวจสอบว่าไฟล์ PDF ถูกสร้างขึ้น
            if os.path.exists(pdf_path):
                print(f"PDF created successfully: {pdf_path}")
                return pdf_path
            else:
                print(f"PDF file not found at expected location: {pdf_path}")
                return excel_path
        else:
            print(f"LibreOffice conversion failed with return code: {result.returncode}")
            print(f"Error: {result.stderr}")
            return excel_path
            
    except subprocess.TimeoutExpired:
        print("LibreOffice conversion timed out")
        return excel_path
    except FileNotFoundError:
        print("LibreOffice not found on system")
        return excel_path
    except Exception as e:
        print(f"Error in LibreOffice conversion: {str(e)}")
        return excel_path

def convert_excel_to_pdf(excel_path):
    """Main PDF conversion function - only uses LibreOffice"""
    return convert_excel_to_pdf_libreoffice(excel_path)

def safe_convert_to_pdf(excel_path):
    """Safe conversion with fallback to Excel if PDF fails"""
    
    try:
        print("Attempting LibreOffice PDF conversion...")
        result = convert_excel_to_pdf_libreoffice(excel_path)
        
        # ถ้าได้ PDF file กลับมา ถือว่าสำเร็จ
        if result.endswith('.pdf') and os.path.exists(result):
            print("PDF conversion successful")
            return result
        else:
            print("PDF conversion failed, returning Excel file")
            return excel_path
                
    except Exception as e:
        print(f"PDF conversion error: {str(e)}, returning Excel file")
        return excel_path


@app.route("/generate_excel_form", methods=["POST"])
def generate_excel_form():
    try:
        booking_no = request.form.get('booking_no')
        booking_type = request.form.get('booking_type', 'tour')
        
        if not booking_no:
            return jsonify({"success": False, "message": "Booking number not specified"}), 400
        
        print(f"Processing Excel generation for {booking_type} booking: {booking_no}")
        
        # เชื่อมต่อฐานข้อมูล
        conn = get_db_connection()
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        
        # ✅ เลือก table และ query ตามประเภท
        if booking_type == 'tour':
            table_name = 'tour_rental'
            
            query = f"""
            SELECT 
                booking_no, booking_date, travel_date, pickup_time, 
                customer_name, customer_surname,
                CONCAT(customer_name, ' ', customer_surname) AS full_name,
                room, payment_status, staff_name, received,
                quantity, detail, payment_method, remark, discount, price
            FROM {table_name} 
            WHERE booking_no = %s
            """
            
        else:  # motorbike
            table_name = 'motorbike_rental'
            
            # ✅ สำหรับ motorbike ต้องดึง company_name และ price ด้วย
            query = f"""
            SELECT 
                booking_no, booking_date, travel_date, pickup_time, 
                customer_name, customer_surname,
                CONCAT(customer_name, ' ', customer_surname) AS full_name,
                room, payment_status, staff_name, received,
                quantity, detail, payment_method, remark, discount,
                company_name, price, start_booking_date, end_booking_date
            FROM {table_name} 
            WHERE booking_no = %s
            """
        
        cursor.execute(query, (booking_no,))
        booking = cursor.fetchone()
        
        if not booking:
            cursor.close()
            conn.close()
            return jsonify({"success": False, "message": f"Booking number not found: {booking_no}"}), 404
        
        # สร้างไฟล์ Excel
        excel_file = create_excel_form(booking, booking_type)
        
        # แปลงเป็น PDF
        pdf_file = convert_excel_to_pdf(excel_file)
        
        cursor.close()
        conn.close()
        
        file_prefix = "Tour" if booking_type == 'tour' else "Motorbike"

        print(f"File prefix: {file_prefix}")
        print(f"Download name: {datetime.now().strftime('%Y%m%d')}_{booking_no}_{file_prefix}_Booking.pdf")
        
        # ตรวจสอบว่าได้ PDF หรือ Excel
        if pdf_file.endswith('.pdf'):
            return send_file(
                pdf_file,
                as_attachment=False,  # เปิดในเบราว์เซอร์
                download_name=f"{datetime.now().strftime('%Y%m%d')}_{booking_no}_{file_prefix}_Booking.pdf",
                mimetype='application/pdf'
            )
        else:
            # ถ้าแปลง PDF ไม่ได้ ส่ง Excel แทน
            return send_file(
                excel_file,
                as_attachment=True,
                download_name=f"{datetime.now().strftime('%Y%m%d')}_{booking_no}_{file_prefix}_Booking.xlsx",
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        
    except Exception as e:
        print(f"General error: {str(e)}")
        return jsonify({"success": False, "message": f"Error: {str(e)}"}), 500



# เพิ่มฟังก์ชันทำความสะอาดไฟล์ชั่วคราว
def cleanup_temp_files():
    """ลบไฟล์ชั่วคราวเก่า"""
    temp_dir = tempfile.gettempdir()
    current_time = datetime.now()
    
    for filename in os.listdir(temp_dir):
        if filename.endswith(('.xlsx', '.pdf')):
            file_path = os.path.join(temp_dir, filename)
            try:
                # ลบไฟล์ที่เก่ากว่า 1 ชั่วโมง
                file_time = datetime.fromtimestamp(os.path.getctime(file_path))
                if (current_time - file_time).seconds > 3600:
                    os.remove(file_path)
                    print(f"Cleaned up temp file: {filename}")
            except Exception as e:
                print(f"Error cleaning temp file {filename}: {e}")

def create_excel_form(booking, booking_type='tour'):  
    """Create Excel form from booking data"""
    # Define template path in static folder
    script_dir = os.path.dirname(os.path.abspath(__file__))
    
    # ✅ เลือก template ตามประเภท
    if booking_type == 'motorbike':
        template_filename = 'motorbike_template.xlsx'
    else:
        template_filename = 'tour_template.xlsx'
    
    template_path = os.path.join(script_dir, 'static', 'templates', template_filename)
    
    # ตรวจสอบว่าไฟล์ template มีอยู่จริง
    if not os.path.exists(template_path):
        print(f"Template not found at: {template_path}")
        
        templates_dir = os.path.join(script_dir, 'static', 'templates')
        if not os.path.exists(templates_dir):
            os.makedirs(templates_dir)
            print(f"Created templates directory: {templates_dir}")
        
        raise FileNotFoundError(f"Not found Excel template at: {template_path}")
    
    # โหลด Excel template
    workbook = openpyxl.load_workbook(template_path)
    sheet = workbook.active
    
    # ✅ แยก logic ตามประเภท booking
    if booking_type == 'motorbike':
        # ✅ สำหรับ Motorbike - รองรับ comma-separated data
        return create_motorbike_excel_form(booking, workbook, sheet, script_dir)
    else:
        # ✅ สำหรับ Tour - ใช้วิธีเดิม
        return create_tour_excel_form(booking, workbook, sheet, script_dir)

def create_tour_excel_form(booking, workbook, sheet, script_dir):
    """สร้าง Excel form สำหรับ Tour (วิธีเดิม)"""
    
    # Cell mappings ตามที่กำหนด (เหมือนเดิม)
    cell_mappings = {
        "booking_date": ["J3", "J30"],
        "booking_no": ["J4", "J31"],
        "customer_name": ["D11", "D38", "G25", "G52"],
        "room": ["D13", "D40"],
        "travel_date": ["F13", "F40"],
        "pickup_time": ["I13", "I40"],
        "payment_status": ["D22", "D49"],
        "staff_name": ["B25", "B52"],
        "detail": ["D16", "D43"],
        "price": ["H16", "H43"],
        "payment_method": ["H22", "H49"],
        "remark": ["D21", "D48"],
        "discount": ["J19", "J46"],
        "received": ["J20", "J47"]
    }
    
    # Formula cells
    formula_cells = {
        "J16": "=H16*I16",  # Total = Price * Quantity for first section
        "J43": "=H43*I43"   # Total = Price * Quantity for second section
    }
    
    # Value mappings
    value_mappings = {
        "quantity": ["I16", "I43"],  # Number of people
        "price": ["H16", "H43"]
    }
    
    # ใส่ข้อมูลลงในเซลล์ตาม mappings
    for field, cells in cell_mappings.items():
        value = None
        
        # กรณีพิเศษสำหรับชื่อลูกค้าเต็ม
        if field == "customer_name":
            value = booking["full_name"]
        # ✅ เพิ่ม capitalize สำหรับ payment_status
        elif field == "payment_status":
            if booking.get(field):
                original_value = booking[field]
                value = str(booking[field]).title()  # ใช้ title() แทน capitalize()
                print(f"DEBUG payment_status: '{original_value}' → '{value}'")
            else:
                value = ""
                print("DEBUG payment_status: No value found")
        elif field in booking and booking[field] is not None:
            value = booking[field]
        
        # จัดรูปแบบวันที่และเวลา
        if field == "booking_date" or field == "travel_date":
            if value:
                if isinstance(value, (date, datetime)):
                    value = value.strftime("%d/%m/%Y")
        elif field == "pickup_time" and value:
            if isinstance(value, (time, datetime)):
                value = value.strftime("%H:%M")
            elif hasattr(value, 'total_seconds'):
                # กรณีเป็น timedelta
                total_seconds = int(value.total_seconds())
                hours = total_seconds // 3600
                minutes = (total_seconds % 3600) // 60
                value = f"{hours:02d}:{minutes:02d}"
            
        # ใส่ข้อมูลลงในทุกเซลล์ที่เกี่ยวข้อง
        if value is not None:
            for cell in cells:
                sheet[cell] = value
    
    # ใส่ข้อมูลจำนวนคน
    if "quantity" in booking and booking["quantity"] is not None:
        for cell in value_mappings["quantity"]:
            sheet[cell] = booking["quantity"]

    if "price" in booking and booking["price"] is not None:
        for cell in value_mappings["price"]:
            sheet[cell] = booking["price"]
    
    # ใส่ข้อมูลราคา
    if "received" in booking and booking["received"] is not None:
            sheet[cell] = booking["received"]
    
    # ใส่สูตรคำนวณ
    for cell, formula in formula_cells.items():
        sheet[cell].value = formula

    # เพิ่มรูปภาพ Tour
    add_tour_images(sheet, script_dir)
    
    # สร้างไฟล์ชั่วคราว
    temp_dir = tempfile.gettempdir()
    temp_file = os.path.join(temp_dir, f"{uuid.uuid4()}.xlsx")
    workbook.save(temp_file)
    
    return temp_file

def create_motorbike_excel_form(booking, workbook, sheet, script_dir):
    """สร้าง Excel form สำหรับ Motorbike แบบแยกบรรทัด"""
     
    basic_mappings = {
        "booking_date": ["I4", "I29"],
        "booking_no": ["I6", "I30"],
        "customer_name": ["C11", "H24", "C35", "H48"],
        "room": ["C12", "C36"],
        "travel_date": ["E12", "E36"],
        "pickup_time": ["H12", "H36"],
        "payment_status": ["C20", "C44"],
        "staff_name": ["B24", "B48"],
        "payment_method": ["D21", "D45"],
        "remark": ["G21", "G45"],
    }
    
    # ใส่ข้อมูลพื้นฐาน
    for field, cells in basic_mappings.items():
        value = None
        
        if field == "customer_name":
            value = booking.get("full_name", "")
        # ✅ เพิ่ม capitalize สำหรับ payment_status
        elif field == "payment_status":
            if booking.get(field):
                original_value = booking[field]
                value = str(booking[field]).title()  # ใช้ title() แทน capitalize()
                print(f"DEBUG payment_status: '{original_value}' → '{value}'")
            else:
                value = ""
                print("DEBUG payment_status: No value found")
        elif field in booking and booking[field] is not None:
            value = booking[field]
        
        # จัดรูปแบบวันที่และเวลา
        if field == "booking_date" or field == "travel_date":
            if value:
                if isinstance(value, (date, datetime)):
                    value = value.strftime("%d/%m/%Y")
        elif field == "pickup_time" and value:
            if isinstance(value, (time, datetime)):
                value = value.strftime("%H:%M")
            elif hasattr(value, 'total_seconds'):
                total_seconds = int(value.total_seconds())
                hours = total_seconds // 3600
                minutes = (total_seconds % 3600) // 60
                value = f"{hours:02d}:{minutes:02d}"
            
        if value is not None:
            for cell in cells:
                sheet[cell] = value
    
    # ✅ แยกข้อมูล comma-separated สำหรับ motorbike (แบบง่าย)
    companies = []
    details = []
    prices = []
    quantities = []
    
    try:
        # ✅ แปลง company_name
        if booking.get('company_name'):
            company_data = str(booking['company_name']).strip()
            if company_data:
                companies = [c.strip() for c in company_data.split(',') if c.strip()]
        
        # ✅ แปลง detail
        if booking.get('detail'):
            detail_data = str(booking['detail']).strip()
            if detail_data:
                details = [d.strip() for d in detail_data.split(',') if d.strip()]
        
        # ✅ แปลง price - ง่ายๆ
        if booking.get('price'):
            price_data = str(booking['price']).strip()
            if price_data:
                prices = [float(p.strip()) for p in price_data.split(',') if p.strip()]
        
        # ✅ แปลง quantity - รองรับทั้ง {1,2,3} และ 1,2,3
        if booking.get('quantity'):
            qty_data = str(booking['quantity']).strip()
            if qty_data:
                # ลบ { } ถ้ามี
                if qty_data.startswith('{') and qty_data.endswith('}'):
                    qty_data = qty_data[1:-1]
                
                quantities = [int(q.strip()) for q in qty_data.split(',') if q.strip()]
        
        # ✅ Debug: แสดงผลการแปลง
        print(f"Parsed companies: {companies}")
        print(f"Parsed details: {details}")
        print(f"Parsed prices: {prices}")
        print(f"Parsed quantities: {quantities}")
        
    except Exception as e:
        print(f"Error parsing motorbike data: {str(e)}")
        # ถ้า error ให้ใช้ข้อมูลเปล่า
        companies = ["Error parsing"]
        details = ["Error parsing"]
        prices = [0]
        quantities = [0]
    
    # ✅ ใส่ข้อมูลลงในเซลล์แบบแยกบรรทัด (แบบง่าย)
    try:
        # หน้าแรก: เริ่ม row 16
        row = 14
        for i in range(len(companies)):
            if i < len(companies):
                sheet[f"C{row}"] = companies[i]
                print(f"Set C{row} = {companies[i]}")
            
            if i < len(details):
                sheet[f"E{row}"] = details[i]
                print(f"Set E{row} = {details[i]}")
            
            if i < len(prices):
                sheet[f"G{row}"] = prices[i]
                print(f"Set G{row} = {prices[i]}")
            
            if i < len(quantities):
                sheet[f"I{row}"] = quantities[i]
                print(f"Set I{row} = {quantities[i]}")
            
            # Sub-total
            sheet[f"J{row}"].value = f"=G{row}*I{row}"
            print(f"Set J{row} = =G{row}*I{row}")
            
            row += 1
            
            # จำกัดไม่เกิน 5 บรรทัด
            if row > 20:
                break
        
        # หน้าสอง: เริ่ม row 43 (ข้อมูลเดียวกัน)
        row = 38
        for i in range(len(companies)):
            if i < len(companies):
                sheet[f"C{row}"] = companies[i]
            
            if i < len(details):
                sheet[f"E{row}"] = details[i]
            
            if i < len(prices):
                sheet[f"G{row}"] = prices[i]
            
            if i < len(quantities):
                sheet[f"I{row}"] = quantities[i]
            
            # Sub-total
            sheet[f"J{row}"].value = f"=G{row}*I{row}"
            
            row += 1
            
            # จำกัดไม่เกิน 5 บรรทัด
            if row > 47:
                break
                
    except Exception as e:
        print(f"Error setting motorbike data in cells: {str(e)}")
    
    # ✅ ใส่ discount และ received
    try:
        # Discount
        if booking.get('discount'):
            discount_val = booking['discount']
            if discount_val and str(discount_val).strip() not in ['', '0']:
                sheet["J19"] = f"-{discount_val}"
                sheet["J43"] = f"-{discount_val}"
                print(f"Set discount: -{discount_val}")
        
        # Received
        if booking.get('received'):
            sheet["J20"] = booking['received']
            sheet["J44"] = booking['received']
            print(f"Set received: {booking['received']}")
    
    except Exception as e:
        print(f"Error setting discount/received: {str(e)}")
    
    # เพิ่มรูปภาพ Motorbike
    add_motorbike_images(sheet, script_dir)
    
    # สร้างไฟล์ชั่วคราว
    temp_dir = tempfile.gettempdir()
    temp_file = os.path.join(temp_dir, f"{uuid.uuid4()}.xlsx")
    workbook.save(temp_file)
    
    print(f"=== DEBUG: Excel file saved to {temp_file} ===")
    
    return temp_file

@app.route("/generate_agreement_excel_form", methods=["POST"])
def generate_agree_excel_form():
    try:
        booking_no = request.form.get('booking_no')
        booking_type = request.form.get('booking_type', 'motorbike')  # ✅ เพิ่ม default
        
        if not booking_no:
            return jsonify({"success": False, "message": "Booking number not specified"}), 400
        
        print(f"Processing Agreement generation for {booking_type} booking: {booking_no}")
        
        # เชื่อมต่อฐานข้อมูล
        conn = get_db_connection()
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        
        # ✅ เลือก table ตาม booking_type
        if booking_type == 'motorbike':
            table_name = 'motorbike_rental'
        else:
            # ถ้าต้องการรองรับ tour ในอนาคต
            table_name = 'tour_rental'
            
        # ✅ Query ข้อมูล booking
        query = f"""
        SELECT 
            booking_no, booking_date, travel_date, pickup_time, 
            customer_name, customer_surname,
            CONCAT(customer_name, ' ', customer_surname) AS full_name,
            room, payment_status, staff_name, received,
            quantity, detail, payment_method, remark, discount,
            company_name, price, start_booking_date, end_booking_date
        FROM {table_name} 
        WHERE booking_no = %s
        """
        
        cursor.execute(query, (booking_no,))
        booking = cursor.fetchone()
        
        if not booking:
            cursor.close()
            conn.close()
            return jsonify({"success": False, "message": f"Booking number not found: {booking_no}"}), 404
        
        # ✅ สร้างไฟล์ Excel Agreement
        excel_file = create_agreement_excel_form(booking)
        
        # แปลงเป็น PDF
        pdf_file = convert_excel_to_pdf(excel_file)
        
        cursor.close()
        conn.close()
        
        file_prefix = "Agreement"
        
        # ตรวจสอบว่าได้ PDF หรือ Excel
        if pdf_file.endswith('.pdf'):
            return send_file(
                pdf_file,
                as_attachment=False,  # เปิดในเบราว์เซอร์
                download_name=f"{datetime.now().strftime('%Y%m%d')}_{booking_no}_{file_prefix}.pdf",
                mimetype='application/pdf'
            )
        else:
            # ถ้าแปลง PDF ไม่ได้ ส่ง Excel แทน
            return send_file(
                excel_file,
                as_attachment=True,
                download_name=f"{datetime.now().strftime('%Y%m%d')}_{booking_no}_{file_prefix}.xlsx",
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        
    except Exception as e:
        print(f"Agreement generation error: {str(e)}")
        return jsonify({"success": False, "message": f"Error: {str(e)}"}), 500

# ✅ แก้ไขฟังก์ชัน create_agreement_excel_form
def create_agreement_excel_form(booking):  
    """สร้าง Excel form สำหรับ Agreement"""
    
    # Define template path in static folder
    script_dir = os.path.dirname(os.path.abspath(__file__))
    template_filename = 'agreement_template.xlsx'
    template_path = os.path.join(script_dir, 'static', 'templates', template_filename)
    
    # ตรวจสอบว่าไฟล์ template มีอยู่จริง
    if not os.path.exists(template_path):
        print(f"Agreement template not found at: {template_path}")
        
        templates_dir = os.path.join(script_dir, 'static', 'templates')
        if not os.path.exists(templates_dir):
            os.makedirs(templates_dir)
            print(f"Created templates directory: {templates_dir}")
        
        raise FileNotFoundError(f"Agreement template not found at: {template_path}")
    
    # โหลด Excel template
    workbook = openpyxl.load_workbook(template_path)
    sheet = workbook.active
    
    # ✅ ใส่ข้อมูลลง template
    return create_agreement_data_excel_form(booking, workbook, sheet, script_dir)

def create_agreement_data_excel_form(booking, workbook, sheet, script_dir):
    """ใส่ข้อมูลลงใน Agreement template"""
     
    # ✅ Cell mappings สำหรับ Agreement
    basic_mappings = {
        "booking_no": ["J10", "U10"],           # Booking Number
        "customer_name": ["D12", "O12"],        # Customer Name
        "start_booking_date": ["D13", "O13"],   # Start Date
        "end_booking_date": ["D14", "O14"],     # End Date
    }
    
    # ใส่ข้อมูลพื้นฐาน
    for field, cells in basic_mappings.items():
        value = None

        if booking.get('received'):
            sheet["J13"] = booking['received']
            sheet["U13"] = booking['received']
        
        if field == "customer_name":
            value = booking.get("full_name", "")
        elif field in booking and booking[field] is not None:
            value = booking[field]
        
        # จัดรูปแบบวันที่
        if field in ["start_booking_date", "end_booking_date"]:
            if value:
                if isinstance(value, (date, datetime)):
                    value = value.strftime("%d/%m/%Y")

        if field == "booking_date" or field == "travel_date":
            if value:
                if isinstance(value, (date, datetime)):
                    value = value.strftime("%d/%m/%Y")
            
        if value is not None:
            for cell in cells:
                sheet[cell] = value
    
    # ✅ เพิ่มรูปภาพ Agreement
    add_agreement_images(sheet, script_dir)
    
    # สร้างไฟล์ชั่วคราว
    temp_dir = tempfile.gettempdir()
    temp_file = os.path.join(temp_dir, f"agreement_{uuid.uuid4()}.xlsx")
    workbook.save(temp_file)
    
    print(f"Agreement Excel file saved to {temp_file}")
    
    return temp_file

def add_tour_images(sheet, script_dir):
    """เพิ่มรูปภาพสำหรับ Tour"""
    logo_path = os.path.join(script_dir, 'static', 'image123', 'logscsoscexcel.png')
    
    if os.path.exists(logo_path):
        try:
            # เพิ่มโลโก้ที่ A1 (หน้าแรก)
            img1 = Image(logo_path)
            img1.width = 100
            img1.height = 80
            img1.anchor = 'B1'
            sheet.add_image(img1)
            
            # เพิ่มโลโก้ที่ A46 (หน้าที่สอง)
            img2 = Image(logo_path)
            img2.width = 100
            img2.height = 80
            img2.anchor = 'B28'
            sheet.add_image(img2)
            print(f"Tour logo added successfully")
        except Exception as e:
            print(f"Error adding tour logo: {str(e)}")
    

def add_motorbike_images(sheet, script_dir):
    """เพิ่มรูปภาพสำหรับ Motorbike"""
    logo_path = os.path.join(script_dir, 'static', 'image', 'logoexcel.png')
    
    if os.path.exists(logo_path):
        try:
            # เพิ่มโลโก้ที่ (หน้าแรก)
            img1 = Image(logo_path)
            img1.width = 100
            img1.height = 80
            img1.anchor = 'B2'
            sheet.add_image(img1)
            
            # เพิ่มโลโก้ที่ (หน้าที่สอง)
            img2 = Image(logo_path)
            img2.width = 100
            img2.height = 80
            img2.anchor = 'B27'
            sheet.add_image(img2)

            print(f"Motorbike logo added successfully")
        except Exception as e:
            print(f"Error adding motorbike logo: {str(e)}")

def add_agreement_images(sheet, script_dir):
    """เพิ่มรูปภาพสำหรับ Agreement"""
    logo_path = os.path.join(script_dir, 'static', 'image', 'logoexcel.png')
    
    if os.path.exists(logo_path):
        try:
            # เพิ่มโลโก้หน้าแรก
            img1 = Image(logo_path)
            img1.width = 100
            img1.height = 80
            img1.anchor = 'B2'
            sheet.add_image(img1)
            
            # เพิ่มโลโก้หน้าที่สอง
            img2 = Image(logo_path)
            img2.width = 100
            img2.height = 80
            img2.anchor = 'M2'
            sheet.add_image(img2)

            print(f"Agreement logos added successfully")
        except Exception as e:
            print(f"Error adding agreement logos: {str(e)}")
    else:
        print(f"Agreement logo not found at: {logo_path}")
    
    # เพิ่มรูปอื่นๆ สำหรับ Motorbike
    # additional_images = [
    #     {'path': os.path.join(script_dir, 'static', 'images', 'signature.png'), 'anchor': 'H25'},
    # ]
    
    # for img_info in additional_images:
    #     if os.path.exists(img_info['path']):
    #         try:
    #             img = Image(img_info['path'])
    #             img.width = 80
    #             img.height = 40
    #             img.anchor = img_info['anchor']
    #             sheet.add_image(img)
    #             print(f"Motorbike additional image added")
    #         except Exception as e:
    #             print(f"Error adding motorbike image: {str(e)}")

@app.route("/export_tour", methods=["POST"])
@login_required
def export_tour():
    try:
        # Get filter parameters
        filter_type = request.form.get('filter_type', '')
        payment_status = request.form.get('payment_status', 'all')
        
        # ⚠️ Base query เปลี่ยนจาก tour_motobike_rental เป็น tour_rental
        query = """
        SELECT 
            tr.booking_date,
            tr.booking_no, 
            tr.travel_date, 
            tr.customer_name, 
            tr.customer_surname, 
            tr.company_name,
            tr.detail,
            tr.staff_name,
            tr.quantity, 
            tr.received,
            tr.price,
            tr.payment_status,
            tr.payment_method,
            tr.pickup_time,
            tr.room,
            tt.paid,
            tr.remark,
            tr.discount
        FROM tour_rental tr
        LEFT JOIN tour_tabel tt ON tr.detail = tt.detail 
                                   AND tr.company_name = tt.company_name
        WHERE 1=1
        """
        
        params = []
        
        # Apply date filters (เหมือนเดิม)
        if filter_type == 'date':
            start_date = request.form.get('start_date', '')
            end_date = request.form.get('end_date', '')
            
            if start_date:
                query += " AND tr.travel_date >= %s"
                params.append(start_date)
            
            if end_date:
                query += " AND tr.travel_date  <= %s"
                params.append(end_date)
                
        elif filter_type == 'month':
            start_month = request.form.get('start_month', '')
            end_month = request.form.get('end_month', '')
            
            if start_month:
                start_date = f"{start_month}-01"
                query += " AND tr.travel_date  >= %s"
                params.append(start_date)
            
            if end_month:
                year, month = map(int, end_month.split('-'))
                if month == 12:
                    next_year = year + 1
                    next_month = 1
                else:
                    next_year = year
                    next_month = month + 1
                
                end_date = f"{next_year}-{next_month:02d}-01"
                query += " AND tr.travel_date  < %s"
                params.append(end_date)
                
        elif filter_type == 'year':
            start_year = request.form.get('start_year', '')
            end_year = request.form.get('end_year', '')
            
            if start_year:
                start_date = f"{start_year}-01-01"
                query += " AND tr.travel_date  >= %s"
                params.append(start_date)
            
            if end_year:
                end_date = f"{int(end_year) + 1}-01-01"
                query += " AND tr.travel_date  < %s"
                params.append(end_date)
        
        # ⚠️ ลบ experience_type filter ออก
        
        # Apply payment status filter
        if payment_status == 'paid':
            query += " AND tr.payment_status = 'paid'"
        elif payment_status == 'unpaid':
            query += " AND tr.payment_status = 'unpaid'"
        
        query += " ORDER BY tr.booking_date DESC, tr.booking_no DESC"
        
        # Execute query (เหมือนเดิม)
        conn = get_db_connection()
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        cursor.execute(query, params)
        bookings = cursor.fetchall()
        cursor.close()
        conn.close()
        
        # Create Excel file
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        
        # ⚠️ เปลี่ยน sheet title
        sheet.title = "Tour Bookings"
        
        # Headers (เหมือนเดิม แต่ลบ Type column)
        # headers = [
        #     "Travel Date", "Time", "Booking Date", "Booking No.", "Name&Surname", 
        #     "Room", "Company Name", "Detail", "Quantity", "Price / Person", 
        #     "Received", "Paid", "Amount", "Staff Name", "Payment Method", "Remark", "Discount"
        # ]

        headers = [
            "Travel Date", "Time", "Booking Date", "Booking No.", "Name&Surname", 
            "Room", "Company Name", "Detail", "Quantity", "Price", "Discount", "Selling Price"
        ]

        if session.get('role') == 'admin':
            headers.extend(["Cost", "Amount"])

        headers.extend([
            "Payment Status", "Staff Name", "Payment Method", "Remark"
        ])
        
        # Style headers (เหมือนเดิม)
        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num)
            cell.value = header
            cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
            cell.fill = openpyxl.styles.PatternFill(
                start_color="366092", end_color="366092", fill_type="solid"
            )
            cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
        
        # Add data rows (ลบ exp_type_display)
        for row_num, booking in enumerate(bookings, 2):
            # Format dates and times (เหมือนเดิม)
            booking_date = booking['booking_date']
            if isinstance(booking_date, (date, datetime)):
                booking_date = booking_date.strftime("%d/%m/%Y")
            
            travel_date = booking['travel_date']
            if isinstance(travel_date, (date, datetime)):
                travel_date = travel_date.strftime("%d/%m/%Y")
            
            pickup_time = ''
            if booking['pickup_time']:
                if hasattr(booking['pickup_time'], 'strftime'):
                    pickup_time = booking['pickup_time'].strftime('%H:%M')
                elif hasattr(booking['pickup_time'], 'total_seconds'):
                    total_seconds = int(booking['pickup_time'].total_seconds())
                    hours = total_seconds // 3600
                    minutes = (total_seconds % 3600) // 60
                    pickup_time = f"{hours:02d}:{minutes:02d}"
                else:
                    pickup_time = str(booking['pickup_time'])
            
            price_per_person = 0
            if booking['quantity'] and booking['quantity'] > 0:
                price_per_person = booking['received'] / booking['quantity']
            
            paid_amount = booking['paid'] if booking['paid'] else 0
            
            if booking['payment_status'].lower().startswith('cancelled'):
                amount = 0  # cancel = 0
            else:
                amount = booking['received'] - paid_amount  # paid หรือ unpaid = received - paid

            full_name = f"{booking['customer_name']} {booking['customer_surname']}"
            
            # ⚠️ Row data ลบ exp_type_display ออก
            row_data = [
                travel_date,                     # Travel Date
                pickup_time,                     # Time
                booking_date,                    # Booking Date
                booking['booking_no'],           # Booking No.
                full_name,                       # Name&Surname
                booking['room'],                 # Room
                booking['company_name'],         # Company Name
                booking['detail'],               # Detail
                booking['quantity'],             # Quantity
                booking['price'],      # Price / Person
                booking['discount'],
                booking['received']             # Received
            ]

            if session.get('role') == 'admin':
                row_data.extend([paid_amount, round(amount, 2)])

            row_data.extend([
                booking['payment_status'],
                booking['staff_name'],            # Staff Name
                booking['payment_method'],                
                booking['remark'],
            ])
            
            for col_num, cell_value in enumerate(row_data, 1):
                cell = sheet.cell(row=row_num, column=col_num)
                cell.value = cell_value
        
        # Add Total row (เหมือนเดิม แต่ปรับ column numbers)
        if bookings and session.get('role') == 'admin':
            total_row = len(bookings) + 2
            sheet.cell(row=total_row, column=11).value = "TOTAL"
            sheet.cell(row=total_row, column=11).font = openpyxl.styles.Font(bold=True)
            
            start_data_row = 2
            end_data_row = total_row - 1
            
            # ปรับ column numbers สำหรับ formulas
            sheet.cell(row=total_row, column=12).value = f"=SUM(L{start_data_row}:L{end_data_row})"  # Received
            sheet.cell(row=total_row, column=12).font = openpyxl.styles.Font(bold=True)
            
            sheet.cell(row=total_row, column=13).value = f"=SUM(M{start_data_row}:M{end_data_row})"  # Paid
            sheet.cell(row=total_row, column=13).font = openpyxl.styles.Font(bold=True)
            
            sheet.cell(row=total_row, column=14).value = f"=SUM(N{start_data_row}:N{end_data_row})"  # Amount
            sheet.cell(row=total_row, column=14).font = openpyxl.styles.Font(bold=True)
        
        # Auto-size columns (เหมือนเดิม)
        for column in sheet.columns:
            max_length = 0
            column_letter = openpyxl.utils.get_column_letter(column[0].column)
            for cell in column:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column_letter].width = adjusted_width
        
        # Create temporary file
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        workbook.save(temp_file.name)
        temp_file.close()
        
        # Send the file
        return send_file(
            temp_file.name,
            as_attachment=True,
            download_name=f"{datetime.now().strftime('%Y%m%d')}_Tour_Export.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        
    except Exception as e:
        print(f"Export error: {str(e)}")
        return jsonify({"success": False, "message": f"Error: {str(e)}"}), 500
    
# -------------------------------- Motorbike Page --------------------------------------------------------------

@app.route("/motorbike_rental")
@login_required
def motorbike_rental():
    # เชื่อมต่อกับฐานข้อมูล
    conn = get_db_connection()
    cursor = conn.cursor()
    
    # ดึงข้อมูลห้องจากตาราง room
    query = "SELECT room FROM room ORDER BY room"
    cursor.execute(query)
    room_results = cursor.fetchall()
    rooms = [room[0] for room in room_results]
    
    # ดึงข้อมูล motorbike companies เท่านั้น
    query_motorbike = "SELECT DISTINCT company_name FROM motorbike_tabel ORDER BY company_name"
    cursor.execute(query_motorbike)
    motorbike_companies = [company[0] for company in cursor.fetchall()]
    
    cursor.close()
    conn.close()
    
    today = date.today().isoformat()  # YYYY-MM-DD
    
    
    return render_template('P3_motorbike_rental.html',
                         room=rooms, 
                         today=today,
                         motorbike_companies=motorbike_companies)
    

@app.route("/submit_motorbike_booking", methods=["POST"])
@login_required
def submit_motorbike_booking():
    try:
        # Connect to the database to get the latest booking number
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Get current year and month (in format YY and MM)
        current_date = datetime.now()
        year_prefix = current_date.strftime("%y")  # 2-digit year (e.g., 25 for 2025)
        month_prefix = current_date.strftime("%m")  # 2-digit month (e.g., 05 for May)
        prefix = f"{year_prefix}{month_prefix}"  # e.g., 2505
        
        # Query to find the highest booking number with the current year-month prefix
        query = """
        SELECT booking_no FROM motorbike_rental 
        WHERE booking_no LIKE %s 
        ORDER BY CAST(SUBSTRING(booking_no, 5) AS INTEGER) DESC 
        LIMIT 1
        """
        cursor.execute(query, (f"{prefix}%",))
        result = cursor.fetchone()
        
        if result:
            last_booking_no = result[0]
            try:
                running_number = int(last_booking_no[4:]) + 1
            except (ValueError, IndexError):
                running_number = 1
        else:
            running_number = 1
        
        max_attempts = 1000
        booking_no = None
        
        for attempt in range(max_attempts):
            temp_booking_no = f"M{prefix}{running_number:05d}"
            
            check_query = "SELECT booking_no FROM motorbike_rental WHERE booking_no = %s"
            cursor.execute(check_query, (temp_booking_no,))
            exists = cursor.fetchone()
            
            if not exists:
                booking_no = temp_booking_no
                break
            else:
                running_number += 1
        
        if not booking_no:
            return jsonify({"success": False, "message": "Unable to generate unique booking number"})

        # Get form data
        customer_name = request.form.get('name', '')
        customer_surname = request.form.get('surname', '')
        room = request.form.get('room', '')
        pickup_time = request.form.get('time', '')
        travel_date = request.form.get('searchDate', '')
        payment_status = request.form.get('status', '')
        staff_name = request.form.get('staffName', '')
        start_booking_date = request.form.get('searchDate', '')
        end_booking_date = request.form.get('searchDateTo', '')
        payment_method = request.form.get('paymentmethod', '')
        remark = request.form.get('remark', '')
        discount = request.form.get('discount', '')
        
        # **ใหม่: รับข้อมูล motorbike แบบ comma-separated**
        companies = request.form.get('company', '')  # comma-separated
        details = request.form.get('detail', '')     # comma-separated
        persons_str = request.form.get('persons', '') # comma-separated
        prices_str = request.form.get('price', '')   # comma-separated
        
        # **ใหม่: ตรวจสอบและแปลงข้อมูล**
        if not all([companies, details, persons_str, prices_str]):
            cursor.close()
            conn.close()
            return jsonify({"success": False, "message": "Missing motorbike data"})
        
        try:
            # แปลง comma-separated strings เป็น lists
            company_list = [c.strip() for c in companies.split(',') if c.strip()]
            detail_list = [d.strip() for d in details.split(',') if d.strip()]
            persons_list = [int(p.strip()) for p in persons_str.split(',') if p.strip()]
            prices_list = [float(p.strip()) for p in prices_str.split(',') if p.strip()]
            
            # ตรวจสอบความยาวของ lists ต้องเท่ากัน
            if not (len(company_list) == len(detail_list) == len(persons_list) == len(prices_list)):
                raise ValueError("Motorbike data arrays length mismatch")
            
            # คำนวณยอดรวม
            subtotal = sum(persons_list[i] * prices_list[i] for i in range(len(company_list)))
            
            # คำนวณส่วนลด
            discount_amount = 0
            if discount:
                discount = discount.strip()
                if discount.endswith('%'):
                    # เปอร์เซ็นต์
                    percentage = float(discount.replace('%', ''))
                    discount_amount = subtotal * (percentage / 100)
                else:
                    # จำนวนเงิน
                    discount_amount = float(discount)
            
            # ยอดสุทธิ
            received = subtotal - discount_amount
            
            # ✅ เก็บ persons รวม (สำหรับ quantity field) แบบ comma-separated string
            total_quantity = sum(persons_list)
            quantity_str = persons_str  # เก็บเป็น "1,2,3" แทน {1,2,3}
            
        except (ValueError, TypeError) as e:
            cursor.close()
            conn.close()
            return jsonify({"success": False, "message": f"Invalid motorbike data format: {str(e)}"})
        
        booking_date = date.today().isoformat()
        
        # Validate required fields
        if not all([customer_name, customer_surname, room, pickup_time, travel_date, 
                   payment_status, staff_name, payment_method]):
            cursor.close()
            conn.close()
            return jsonify({"success": False, "message": "Please fill in all required fields"})
        
        # ✅ **แก้ไข: บันทึกข้อมูลแบบ comma-separated string**
        query = """
        INSERT INTO motorbike_rental (
            travel_date, pickup_time, booking_date, booking_no, 
            customer_name, customer_surname, room, company_name, 
            detail, quantity, received, payment_status, 
            staff_name, start_booking_date, end_booking_date, 
            payment_method, remark, discount, price
        ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """
        
        values = (
            travel_date, pickup_time, booking_date, booking_no,
            customer_name, customer_surname, room, 
            companies,      # comma-separated companies: "Company A,Company B"
            details,        # comma-separated details: "Detail A,Detail B"
            quantity_str,   # ✅ comma-separated quantities: "1,2,3" (ไม่ใช่ {1,2,3})
            received,       # total amount after discount
            payment_status, staff_name, start_booking_date, end_booking_date, 
            payment_method, remark, discount, 
            prices_str      # comma-separated prices: "100.0,200.0"
        )
        
        cursor.execute(query, values)
        conn.commit()
        
        cursor.close()
        conn.close()
        
        return jsonify({
            "success": True, 
            "message": "Motorbike booking successfully submitted", 
            "booking_no": booking_no,
            "total_amount": received,
            "items_count": len(company_list)
        })
    
    except Exception as e:
        try:
            if 'cursor' in locals():
                cursor.close()
            if 'conn' in locals():
                conn.close()
        except:
            pass
        
        return jsonify({"success": False, "message": f"Error: {str(e)}"})
    
@app.route("/motorbike_rental_form")
@login_required
def view_motorbike_bookings():
    try:
        # เชื่อมต่อกับฐานข้อมูล
        conn = get_db_connection()
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        
        # ดึงข้อมูลจากตาราง motorbike_rental เท่านั้น
        query = """
        SELECT 
            *
        FROM motorbike_rental
        WHERE travel_date >= CURRENT_DATE
        ORDER BY travel_date ASC
        LIMIT 50
        """
        
        cursor.execute(query)
        bookings = cursor.fetchall()
        
        # ⚠️ เพิ่มส่วนนี้ - ดึง booking_list สำหรับ dropdown
        cursor.execute("SELECT DISTINCT booking_no, customer_name, customer_surname FROM motorbike_rental ORDER BY booking_no DESC")
        booking_list = cursor.fetchall()
        
        # แปลงรูปแบบวันที่และเวลา (เหมือนเดิม)
        for booking in bookings:
            try:
                if booking['company_name'] and ',' in booking['company_name']:
                    companies = booking['company_name'].split(',')
                    details = booking['detail'].split(',') if booking['detail'] else []
                    prices = booking['price'].split(',') if booking['price'] else []
                    
                    # สร้างข้อความแสดงรายการ
                    items = []
                    for i, company in enumerate(companies):
                        detail = details[i] if i < len(details) else ''
                        price = prices[i] if i < len(prices) else ''
                        items.append(f"{company.strip()}-{detail.strip()}")
                    
                    booking['company_display'] = ' | '.join(items)
                    booking['detail_display'] = f"{len(companies)} items"
                else:
                    booking['company_display'] = booking['company_name']
                    booking['detail_display'] = booking['detail']
            except:
                booking['company_display'] = booking['company_name']
                booking['detail_display'] = booking['detail']

            if booking['travel_date']:
                booking['travel_date'] = booking['travel_date'].strftime('%d/%m/%Y')
            if booking['booking_date']:
                booking['booking_date'] = booking['booking_date'].strftime('%d/%m/%Y')
            if booking['start_booking_date'] and booking['start_booking_date']:
                booking['start_booking_date'] = booking['start_booking_date'].strftime('%d/%m/%Y')
            if booking['end_booking_date'] and booking['end_booking_date']:
                booking['end_booking_date'] = booking['end_booking_date'].strftime('%d/%m/%Y')
            
            
            # แปลงเวลา
            if booking['pickup_time']:
                if hasattr(booking['pickup_time'], 'strftime'):
                    booking['pickup_time'] = booking['pickup_time'].strftime('%H:%M')
                else:
                    total_seconds = int(booking['pickup_time'].total_seconds())
                    hours = total_seconds // 3600
                    minutes = (total_seconds % 3600) // 60
                    booking['pickup_time'] = f"{hours:02d}:{minutes:02d}"
        
        cursor.close()
        conn.close()
        
        # ⚠️ เพิ่ม booking_list ในการ return
        return render_template('P3_motorbike_form.html', 
                             bookings=bookings,
                             booking_list=booking_list)
    
    except Exception as e:
        return f"Error: {str(e)}"
        

# ---- อัพเดต Search Bookings ----
@app.route("/search_motorbike_bookings", methods=["GET", "POST"])
@login_required
def search_motorbike_bookings():
    try:
        if request.method == "POST":
            # รับค่าจากฟอร์มค้นหา
            start_date = request.form.get('start_date', '')
            end_date = request.form.get('end_date', '')
            booking_no = request.form.get('booking_no', '')
            name_surname = request.form.get('name_surname', '')
            
            # เชื่อมต่อกับฐานข้อมูล
            conn = get_db_connection()
            cursor = conn.cursor(cursor_factory=RealDictCursor)
            
            query = """
            SELECT 
                booking_no, 
                travel_date, 
                pickup_time,
                customer_name, 
                customer_surname, 
                room, 
                company_name, 
                detail, 
                quantity, 
                received, 
                payment_status, 
                staff_name, 
                start_booking_date, 
                end_booking_date,
                booking_date,
                payment_method,
                remark,
                discount,
                price
            FROM motorbike_rental
            WHERE 1=1
            """
            
            params = []
            
            # เพิ่มเงื่อนไขการค้นหา
            if start_date:
                query += " AND travel_date >= %s"
                params.append(start_date)
            
            if end_date:
                query += " AND travel_date <= %s"
                params.append(end_date)
            
            if booking_no:
                query += " AND booking_no = %s"
                params.append(booking_no)
                
            if name_surname:
                # แยกชื่อและนามสกุล
                name_parts = name_surname.strip().split(' ', 1)
                if len(name_parts) == 2:
                    first_name, last_name = name_parts
                    query += " AND customer_name = %s AND customer_surname = %s"
                    params.extend([first_name, last_name])
                else:
                    # ถ้าไม่สามารถแยกได้ ให้ค้นหาจากชื่อเต็ม
                    query += " AND CONCAT(customer_name, ' ', customer_surname) = %s"
                    params.append(name_surname)
            
            # ⚠️ ลบ experience_type filter ออก
            
            query += " ORDER BY travel_date DESC, booking_no DESC"
            
            cursor.execute(query, params)
            bookings = cursor.fetchall()
            
            # แปลงรูปแบบวันที่และเวลา (เหมือนเดิม)
            for booking in bookings:
                if booking['travel_date']:
                    booking['travel_date'] = booking['travel_date'].strftime('%d/%m/%Y')
                if booking['booking_date']:
                    booking['booking_date'] = booking['booking_date'].strftime('%d/%m/%Y')
                if booking['start_booking_date']:
                    booking['start_booking_date'] = booking['start_booking_date'].strftime('%d/%m/%Y')
                if booking['end_booking_date']:
                    booking['end_booking_date'] = booking['end_booking_date'].strftime('%d/%m/%Y')
                
                if booking['pickup_time']:
                    if hasattr(booking['pickup_time'], 'strftime'):
                        booking['pickup_time'] = booking['pickup_time'].strftime('%H:%M')
                    else:
                        total_seconds = int(booking['pickup_time'].total_seconds())
                        hours = total_seconds // 3600
                        minutes = (total_seconds % 3600) // 60
                        booking['pickup_time'] = f"{hours:02d}:{minutes:02d}"
            
            # ดึง booking list สำหรับ dropdown
            cursor.execute("SELECT DISTINCT booking_no, customer_name, customer_surname FROM motorbike_rental ORDER BY booking_no DESC")
            booking_list = cursor.fetchall()
            
            cursor.close()
            conn.close()
            
            # ดึงข้อมูล room list สำหรับ dropdown
            conn = get_db_connection()
            cursor = conn.cursor()
            query = "SELECT room FROM room ORDER BY room"
            cursor.execute(query)
            room_results = cursor.fetchall()
            room_list = [room[0] for room in room_results]
            cursor.close()
            conn.close()
            
            return render_template('P3_motorbike_form.html', 
                                 bookings=bookings,
                                 booking_list=booking_list,
                                 search_start_date=start_date, 
                                 search_end_date=end_date, 
                                 search_booking_no=booking_no,
                                 search_name_surname=name_surname,
                                 room_list=room_list)
        else:
            # ถ้าเป็น GET request ให้ redirect ไปที่หน้าแสดงการจองทั้งหมด
            return redirect(url_for('view_motorbike_bookings'))
    
    except Exception as e:
        return f"Error: {str(e)}"
    
@app.route("/update_motorbike_booking", methods=["POST"])
@login_required
def update_motorbike_booking():
    try:
        # รับค่าจากฟอร์มแก้ไข
        booking_no = request.form.get('booking_no')
        
        # ข้อมูลพื้นฐาน
        customer_name = request.form.get('name', '')
        customer_surname = request.form.get('surname', '')
        room = request.form.get('room', '')
        pickup_time = request.form.get('time', '')
        travel_date = request.form.get('searchDate', '')  # เปลี่ยนจาก date เป็น searchDate
        start_booking_date = request.form.get('searchDate', '')  # ใช้ travel_date เป็น start_booking_date
        end_booking_date = request.form.get('searchDateTo', '')
        payment_status = request.form.get('status', '')
        payment_method = request.form.get('paymentmethod', '')
        staff_name = request.form.get('staffName', '')
        remark = request.form.get('remark', '')
        discount = request.form.get('discount', '')
        
        # ข้อมูล motorbike แบบ comma-separated
        companies = request.form.get('company', '')  # comma-separated
        details = request.form.get('detail', '')     # comma-separated
        persons_str = request.form.get('persons', '') # comma-separated
        prices_str = request.form.get('price', '')   # comma-separated
        
        if not booking_no:
            return jsonify({"success": False, "message": "Booking number is required"})
        
        # ตรวจสอบข้อมูล motorbike
        if not all([companies, details, persons_str, prices_str]):
            return jsonify({"success": False, "message": "Missing motorbike data"})
        
        try:
            # แปลง comma-separated strings เป็น lists เพื่อตรวจสอบ
            company_list = [c.strip() for c in companies.split(',') if c.strip()]
            detail_list = [d.strip() for d in details.split(',') if d.strip()]
            persons_list = [int(p.strip()) for p in persons_str.split(',') if p.strip()]
            prices_list = [float(p.strip()) for p in prices_str.split(',') if p.strip()]
            
            # ตรวจสอบความยาวของ lists ต้องเท่ากัน
            if not (len(company_list) == len(detail_list) == len(persons_list) == len(prices_list)):
                raise ValueError("Motorbike data arrays length mismatch")
            
            # คำนวณยอดรวม
            subtotal = sum(persons_list[i] * prices_list[i] for i in range(len(company_list)))
            
            # คำนวณส่วนลด
            discount_amount = 0
            if discount:
                discount = discount.strip()
                if discount.endswith('%'):
                    # เปอร์เซ็นต์
                    percentage = float(discount.replace('%', ''))
                    discount_amount = subtotal * (percentage / 100)
                else:
                    # จำนวนเงิน
                    discount_amount = float(discount)
            
            # ยอดสุทธิ
            received = subtotal - discount_amount
            
            # เก็บ quantity เป็น comma-separated string
            quantity_str = persons_str  # เก็บเป็น "1,2,3"
            
        except (ValueError, TypeError) as e:
            return jsonify({"success": False, "message": f"Invalid motorbike data format: {str(e)}"})
        
        # เชื่อมต่อกับฐานข้อมูล
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # อัปเดตข้อมูลในฐานข้อมูล
        query = """
        UPDATE motorbike_rental SET
            travel_date = %s,
            pickup_time = %s,
            customer_name = %s,
            customer_surname = %s,
            room = %s,
            company_name = %s,
            detail = %s,
            quantity = %s,
            received = %s,
            payment_status = %s,
            payment_method = %s,
            staff_name = %s,
            start_booking_date = %s,
            end_booking_date = %s,
            remark = %s,
            discount = %s,
            price = %s
        WHERE booking_no = %s
        """
        
        values = (
            travel_date, pickup_time, customer_name, customer_surname,
            room, 
            companies,      # comma-separated companies: "Company A,Company B"
            details,        # comma-separated details: "Detail A,Detail B"
            quantity_str,   # comma-separated quantities: "1,2,3"
            received,       # total amount after discount
            payment_status, payment_method, staff_name,
            start_booking_date, end_booking_date, remark, discount,
            prices_str,     # comma-separated prices: "100.0,200.0"
            booking_no
        )
        
        cursor.execute(query, values)
        conn.commit()
        
        updated = cursor.rowcount > 0
        
        cursor.close()
        conn.close()
        
        if updated:
            return jsonify({
                "success": True, 
                "message": "Motorbike booking successfully updated",
                "total_amount": received,
                "items_count": len(company_list)
            })
        else:
            return jsonify({"success": False, "message": "No changes were made or booking not found"})
        
    except Exception as e:
        return jsonify({"success": False, "message": f"Error: {str(e)}"})

@app.route("/export_motorbike", methods=["POST"])
@login_required
def export_motorbike():
    try:
        # Get filter parameters
        filter_type = request.form.get('filter_type', '')
        payment_status = request.form.get('payment_status', 'all')
        
        query = """
        SELECT 
            booking_date,
            booking_no, 
            travel_date, 
            customer_name, 
            customer_surname, 
            company_name,
            detail,
            staff_name,
            quantity,
            received,
            payment_status,
            pickup_time,
            room,
            payment_method,
            remark,
            discount,
            price
        FROM motorbike_rental
        WHERE 1=1
        """
        
        params = []

        total_paid = 0
        
        # Apply date filters
        if filter_type == 'date':
            start_date = request.form.get('start_date', '')
            end_date = request.form.get('end_date', '')
            
            if start_date:
                query += " AND travel_date >= %s"
                params.append(start_date)
            
            if end_date:
                query += " AND travel_date <= %s"
                params.append(end_date)
                
        elif filter_type == 'month':
            start_month = request.form.get('start_month', '')
            end_month = request.form.get('end_month', '')
            
            if start_month:
                start_date = f"{start_month}-01"
                query += " AND travel_date >= %s"
                params.append(start_date)
            
            if end_month:
                year, month = map(int, end_month.split('-'))
                if month == 12:
                    next_year = year + 1
                    next_month = 1
                else:
                    next_year = year
                    next_month = month + 1
                
                end_date = f"{next_year}-{next_month:02d}-01"
                query += " AND travel_date < %s"
                params.append(end_date)
                
        elif filter_type == 'year':
            start_year = request.form.get('start_year', '')
            end_year = request.form.get('end_year', '')
            
            if start_year:
                start_date = f"{start_year}-01-01"
                query += " AND travel_date >= %s"
                params.append(start_date)
            
            if end_year:
                end_date = f"{int(end_year) + 1}-01-01"
                query += " AND travel_date < %s"
                params.append(end_date)
        
        # Apply payment status filter
        if payment_status == 'paid':
            query += " AND payment_status = 'paid'"
        elif payment_status == 'unpaid':
            query += " AND payment_status = 'unpaid'"
        
        query += " ORDER BY booking_date DESC, booking_no DESC"
        
        # Execute query
        conn = get_db_connection()
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        cursor.execute(query, params)
        bookings = cursor.fetchall()
        cursor.close()
        conn.close()
        
        # Create Excel file
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Motorbike Bookings"
        
        # Headers
        headers = [
            "Travel Date", "Time", "Booking Date", "Booking No.", "Name&Surname", 
            "Room", "Company Name", "Detail", "Quantity", "Price", "Discount", "Selling Price"
        ]

        if session.get('role') == 'admin':
            headers.extend(["Cost", "Amount"])

        headers.extend([
            "Payment Status", "Staff Name", "Payment Method", "Remark"
        ])
        
        # Style headers
        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num)
            cell.value = header
            cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
            cell.fill = openpyxl.styles.PatternFill(
                start_color="366092", end_color="366092", fill_type="solid"
            )
            cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
        
        # Add data rows
        for row_num, booking in enumerate(bookings, 2):
            # Format dates and times
            booking_date = booking['booking_date']
            if isinstance(booking_date, (date, datetime)):
                booking_date = booking_date.strftime("%d/%m/%Y")
            
            travel_date = booking['travel_date']
            if isinstance(travel_date, (date, datetime)):
                travel_date = travel_date.strftime("%d/%m/%Y")
            
            pickup_time = ''
            if booking['pickup_time']:
                if hasattr(booking['pickup_time'], 'strftime'):
                    pickup_time = booking['pickup_time'].strftime('%H:%M')
                elif hasattr(booking['pickup_time'], 'total_seconds'):
                    total_seconds = int(booking['pickup_time'].total_seconds())
                    hours = total_seconds // 3600
                    minutes = (total_seconds % 3600) // 60
                    pickup_time = f"{hours:02d}:{minutes:02d}"
                else:
                    pickup_time = str(booking['pickup_time'])
            
            # Get data from booking record and format comma-separated values with line breaks
            quantities_str = booking['quantity'] if booking['quantity'] else ""
            prices_str = booking['price'] if booking['price'] else ""
            company_name_str = booking['company_name'] if booking['company_name'] else ""
            detail_str = booking['detail'] if booking['detail'] else ""
            received = booking['received'] if booking['received'] else 0
            
            # Convert comma-separated values to line-separated for Excel display
            def format_comma_to_lines(text):
                if not text:
                    return ""
                return '\n'.join([item.strip() for item in str(text).split(',') if item.strip()])
            
            company_name_str = booking['company_name'] or ""
            detail_str = booking['detail'] or ""

            # แปลงเป็น list
            company_list = [c.strip() for c in company_name_str.split(',') if c.strip()]
            detail_list = [d.strip() for d in detail_str.split(',') if d.strip()]

            # ดึงข้อมูล paid จาก motorbike_tabel
            conn = get_db_connection()
            cursor = conn.cursor()
            cursor.execute("SELECT company_name, detail, paid FROM motorbike_tabel")
            rows = cursor.fetchall()
            cursor.close()
            conn.close()

            # สร้าง dict โดยใช้ (company_name, detail) เป็น key
            paid_lookup = {(row[0], row[1]): row[2] for row in rows}

            # แปลง quantity เป็น list ของ float
            quantity_list = [float(q.strip()) if q.strip().replace('.', '', 1).isdigit() else 0 for q in quantities_str.split(',') if q.strip()]

            # สร้าง paid_list โดย match ทีละคู่และคูณกับ quantity
            paid_list = []
            for i, (cname, dval) in enumerate(zip(company_list, detail_list)):
                paid_value = float(paid_lookup.get((cname, dval), 0) or 0)
                qty = quantity_list[i] if i < len(quantity_list) else 1
                paid_amount = paid_value * qty
                paid_list.append(f"{paid_amount:.2f}")

            # รวมยอด paid ในแต่ละแถวสำหรับรวม total ทีหลัง
            paid_total = sum(float(p) for p in paid_list)

            # เก็บเป็นหลายบรรทัดสำหรับใส่ใน Excel
            paid_formatted = '\n'.join(paid_list)

            # สะสมยอดรวมเพื่อใช้ในแถว TOTAL
            total_paid += paid_total

            
            company_name_formatted = format_comma_to_lines(company_name_str)
            detail_formatted = format_comma_to_lines(detail_str)
            quantities_formatted = format_comma_to_lines(quantities_str)
            prices_formatted = format_comma_to_lines(prices_str)
            
            full_name = f"{booking['customer_name'] or ''} {booking['customer_surname'] or ''}".strip()
            
            row_data = [
                travel_date,                     # Travel Date
                pickup_time,                     # Time
                booking_date,                    # Booking Date
                booking['booking_no'],           # Booking No.
                full_name,                       # Name&Surname
                booking['room'],                 # Room
                company_name_formatted,          # Company Name (line-separated)
                detail_formatted,                # Detail (line-separated)
                quantities_formatted,            # Quantity (line-separated)
                prices_formatted,                # Price (line-separated)
                booking['discount'],              # Discount
                received                        # Received
            ]

            if session.get('role') == 'admin':

                # คำนวณ amount = received - sum(paid_list)
                try:
                    received_float = float(booking['received']) if booking['received'] else 0
                    paid_total = sum(float(p) for p in paid_list)
                    if booking['payment_status'].lower().startswith('cancelled'):
                        amount = 0  # cancel = 0
                    else:
                        amount = round(received_float - paid_total, 2)  # paid หรือ unpaid = received - paid
                except Exception:
                    amount = 0

                row_data.extend([paid_formatted, amount])


            row_data.extend([
                booking['payment_status'],       # Payment Status
                booking['staff_name'],           # Staff Name
                booking['payment_method'],       # Payment Method
                booking['remark'],               # Remark
            ])
            
            for col_num, cell_value in enumerate(row_data, 1):
                cell = sheet.cell(row=row_num, column=col_num)
                cell.value = cell_value
                
                # Enable text wrapping for cells with line breaks
                if cell_value and isinstance(cell_value, str) and '\n' in cell_value:
                    cell.alignment = openpyxl.styles.Alignment(wrap_text=True, vertical='top')
        
        # Add Total row
        if bookings and session.get('role') == 'admin':
            total_row = len(bookings) + 2
            sheet.cell(row=total_row, column=11).value = "TOTAL"  # Price column
            sheet.cell(row=total_row, column=11).font = openpyxl.styles.Font(bold=True)
            
            start_data_row = 2
            end_data_row = total_row - 1
            
            # Add formula for total received
            sheet.cell(row=total_row, column=12).value = f"=SUM(L{start_data_row}:L{end_data_row})"  # Received
            sheet.cell(row=total_row, column=12).font = openpyxl.styles.Font(bold=True)

            # Add formula for total paid (column 13)
            sheet.cell(row=total_row, column=13).value = total_paid
            sheet.cell(row=total_row, column=13).font = openpyxl.styles.Font(bold=True)

            # Add formula for total amount (column 14)
            sheet.cell(row=total_row, column=14).value = f"=SUM(N{start_data_row}:N{end_data_row})"
            sheet.cell(row=total_row, column=14).font = openpyxl.styles.Font(bold=True)
        
        # Auto-size columns and set row heights for wrapped text
        for column in sheet.columns:
            max_length = 0
            column_letter = openpyxl.utils.get_column_letter(column[0].column)
            for cell in column:
                if cell.value:
                    # Count lines for height calculation
                    if isinstance(cell.value, str) and '\n' in cell.value:
                        line_count = cell.value.count('\n') + 1
                        # Set row height based on number of lines
                        sheet.row_dimensions[cell.row].height = max(
                            sheet.row_dimensions[cell.row].height or 15, 
                            15 * line_count
                        )
                    
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column_letter].width = adjusted_width
        
        # Create temporary file
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        workbook.save(temp_file.name)
        temp_file.close()
        
        # Send the file
        return send_file(
            temp_file.name,
            as_attachment=True,
            download_name=f"{datetime.now().strftime('%Y%m%d')}_Motorbike_Export.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        
    except Exception as e:
        print(f"Export error: {str(e)}")
        return jsonify({"success": False, "message": f"Error: {str(e)}"}), 500

# -------------------------------- home transfer Page --------------------------------------------------------------
# app.py
# เพิ่ม API endpoint นี้ใน Flask app ของคุณ

@app.route("/api/get_transfer_routes", methods=["GET"])
@login_required
def get_transfer_routes():
    """API endpoint สำหรับดึงข้อมูล transfer routes จาก database"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        
        # ดึงข้อมูลทั้งหมดจาก transfer_tabel
        query = """
        SELECT place_from, place_to, passengers, received, paid 
        FROM transfer_tabel 
        ORDER BY place_from, place_to, passengers
        """
        cursor.execute(query)
        routes = cursor.fetchall()
        
        cursor.close()
        conn.close()
        
        # แปลงข้อมูลให้เป็น format ที่ JavaScript ใช้ได้
        transfer_routes = []
        for route in routes:
            transfer_routes.append({
                'place_from': route['place_from'],
                'place_to': route['place_to'],
                'passengers': int(route['passengers']),  # แปลงเป็น integer
                'received': float(route['received']) if route['received'] else 0,
                'paid': float(route['paid']) if route['paid'] else 0
            })
        
        return jsonify({
            "success": True, 
            "routes": transfer_routes,
            "count": len(transfer_routes)
        })
        
    except Exception as e:
        return jsonify({
            "success": False, 
            "message": f"Error: {str(e)}",
            "routes": [],
            "count": 0
        })

# ปรับปรุง home_transfer route เพื่อส่งข้อมูลที่ง่ายขึ้น
@app.route("/home_transfer")
@login_required
def home_transfer():
    try:
        # Connect to database to get transfer locations
        conn = get_db_connection()
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        
        # Get unique destinations for departure transfers (where place_from is resort)
        departure_query = """
        SELECT DISTINCT place_to as destination
        FROM transfer_tabel 
        WHERE place_from = 'Lamai Bayview Boutique Resort'
        ORDER BY place_to
        """
        cursor.execute(departure_query)
        departure_destinations = cursor.fetchall()
        
        # Get unique origins for arrival transfers (where place_to is resort)
        arrival_query = """
        SELECT DISTINCT place_from as origin
        FROM transfer_tabel 
        WHERE place_to = 'Lamai Bayview Boutique Resort'
        ORDER BY place_from
        """
        cursor.execute(arrival_query)
        arrival_origins = cursor.fetchall()
        
        cursor.close()
        conn.close()
        
        return render_template("P2_home_transfer.html", 
                             departure_destinations=departure_destinations,
                             arrival_origins=arrival_origins)
    
    except Exception as e:
        print(f"Error loading transfer data: {str(e)}")
        return render_template("P2_home_transfer.html", 
                             departure_destinations=[],
                             arrival_origins=[])

# ลบ API endpoint get_transfer_price เดิมออก เพราะไม่ใช้แล้ว
# ใช้ข้อมูลจาก /api/get_transfer_routes แทน

@app.route("/submit_transfer_booking", methods=["POST"])
@login_required
def submit_transfer_booking():
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # ✅ แก้ไขการสร้าง prefix
        current_date = datetime.now()
        year_prefix = current_date.strftime("%y")   # 25
        month_prefix = current_date.strftime("%m")  # 12 (ธันวาคม)
        prefix = f"H{year_prefix}{month_prefix}"    # H2512
        
        print(f"Current date: {current_date}")  # Debug
        print(f"Generated prefix: {prefix}")    # Debug
        
        # ✅ แก้ไข query หา running number
        query = """
        SELECT booking_no FROM transfer_rental 
        WHERE booking_no LIKE %s 
        ORDER BY CAST(SUBSTRING(booking_no FROM 6) AS INTEGER) DESC 
        LIMIT 1
        """
        cursor.execute(query, (f"{prefix}%",))
        result = cursor.fetchone()
        
        if result:
            last_booking_no = result[0]
            print(f"Last booking: {last_booking_no}")  # Debug
            
            # ✅ ดึง running number ตัวสุดท้าย
            try:
                # ดึงตัวเลขหลัง prefix (ตำแหน่งที่ 5 เป็นต้นไป)
                running_number = int(last_booking_no[5:]) + 1
            except (ValueError, IndexError):
                running_number = 1
        else:
            running_number = 1
        
        print(f"Next running number: {running_number}")  # Debug
        
        # ✅ สร้าง booking number ที่ถูกต้อง
        booking_no = f"{prefix}{running_number:05d}"  # 5 หลัก
        print(f"Generated booking_no: {booking_no}")   # Debug
        
        # ✅ ตรวจสอบว่าไม่ซ้ำ
        check_query = "SELECT booking_no FROM transfer_rental WHERE booking_no = %s"
        cursor.execute(check_query, (booking_no,))
        exists = cursor.fetchone()
        
        if exists:
            # ถ้าซ้ำให้เพิ่ม running number
            running_number += 1
            booking_no = f"{prefix}{running_number:05d}"
            print(f"Conflict detected, new booking_no: {booking_no}")  # Debug
        
        # ส่วนที่เหลือเหมือนเดิม...
        # Get form data
        customer_name = request.form.get('name', '')
        customer_surname = request.form.get('surname', '')
        place_from = request.form.get('place_from', '')
        place_to = request.form.get('place_to', '')
        transfer_type = request.form.get('transferType', '')
        detail = request.form.get('detail', '')
        pickup_time = request.form.get('time', '')
        quantity = request.form.get('persons', '0')
        received = request.form.get('received', '0')
        payment_status = request.form.get('status', '')
        staff_name = request.form.get('staffName', '')
        driver_name = request.form.get('driverName', '')
        price = request.form.get('price', '0')
        payment_method = request.form.get('paymentmethod', '')
        remark = request.form.get('remark', '')
        discount = request.form.get('discount', '')
        
        booking_date = date.today().isoformat()
        travel_date = request.form.get('travel_date', date.today().isoformat())
        
        # Insert ข้อมูล
        query = """
        INSERT INTO transfer_rental (
            travel_date, pickup_time, booking_date, booking_no, 
            customer_name, customer_surname, place_from, place_to, 
            transfer_type, detail, quantity, received,
            payment_status, staff_name, driver_name, price, payment_method, remark, discount
        ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """
        
        values = (
            travel_date, pickup_time, booking_date, booking_no,
            customer_name, customer_surname, place_from, place_to,
            transfer_type, detail, quantity, received,
            payment_status, staff_name, driver_name, price, payment_method, remark, discount
        )
        
        cursor.execute(query, values)
        conn.commit()
        
        cursor.close()
        conn.close()
        
        return jsonify({"success": True, "message": "Transfer booking successfully submitted", "booking_no": booking_no})
    
    except Exception as e:
        print(f"Error in submit_transfer_booking: {str(e)}")  # Debug
        return jsonify({"success": False, "message": f"Error: {str(e)}"})

# ✅ แก้ไข route home_transfer_form()
@app.route("/home_transfer_form")
@login_required
def home_transfer_form():
    try:
        # Connect to database
        conn = get_db_connection()
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        
        # Retrieve all bookings from transfer_rental table
        query = """
        SELECT 
            *
        FROM transfer_rental
        WHERE travel_date >= CURRENT_DATE
        ORDER BY travel_date ASC
        LIMIT 50
        """
        
        cursor.execute(query)
        bookings = cursor.fetchall()
        
        # ✅ เพิ่มส่วนนี้ - ดึง booking_list สำหรับ dropdown
        cursor.execute("SELECT DISTINCT booking_no, customer_name, customer_surname FROM transfer_rental ORDER BY booking_no DESC")
        booking_list = cursor.fetchall()
        
        # Format dates and times
        for booking in bookings:
            if booking['travel_date']:
                booking['travel_date'] = booking['travel_date'].strftime('%d/%m/%Y')
            if booking['booking_date']:
                booking['booking_date'] = booking['booking_date'].strftime('%d/%m/%Y')
            
            # Handle pickup_time formatting
            if booking['pickup_time']:
                if hasattr(booking['pickup_time'], 'strftime'):
                    booking['pickup_time'] = booking['pickup_time'].strftime('%H:%M')
                else:
                    # If it's a timedelta
                    total_seconds = int(booking['pickup_time'].total_seconds())
                    hours = total_seconds // 3600
                    minutes = (total_seconds % 3600) // 60
                    booking['pickup_time'] = f"{hours:02d}:{minutes:02d}"
        
        # ดึงข้อมูลจาก transfer_tabel สำหรับ dropdown ใน edit modal
        # Get all unique destinations for departure transfers (where place_from is resort)
        departure_query = """
        SELECT DISTINCT place_to as destination, passengers, received, paid 
        FROM transfer_tabel 
        WHERE place_from = 'Lamai Bayview Boutique Resort'
        ORDER BY place_to
        """
        cursor.execute(departure_query)
        departure_destinations = cursor.fetchall()
        
        # Get all unique origins for arrival transfers (where place_to is resort)
        arrival_query = """
        SELECT DISTINCT place_from as origin, passengers, received, paid 
        FROM transfer_tabel 
        WHERE place_to = 'Lamai Bayview Boutique Resort'
        ORDER BY place_from
        """
        cursor.execute(arrival_query)
        arrival_origins = cursor.fetchall()
        
        cursor.close()
        conn.close()
        
        # ✅ เพิ่ม booking_list ในการ return
        return render_template('P2_home_transfer_form.html', 
                             transfers=bookings,
                             booking_list=booking_list,  # ✅ เพิ่มบรรทัดนี้
                             departure_destinations=departure_destinations,
                             arrival_origins=arrival_origins)
    
    except Exception as e:
        return f"Error: {str(e)}"


# ✅ แก้ไข route search_transfer_bookings() ที่มีอยู่แล้ว
@app.route("/search_transfer_bookings", methods=["GET", "POST"])
@login_required
def search_transfer_bookings():
    try:
        # ✅ เพิ่มส่วนนี้ - ดึง booking_list สำหรับ dropdown (ใช้ทั้ง GET และ POST)
        conn = get_db_connection()
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        
        # ดึง booking list สำหรับ dropdown
        cursor.execute("SELECT DISTINCT booking_no, customer_name, customer_surname FROM transfer_rental ORDER BY booking_no DESC")
        booking_list = cursor.fetchall()
        
        cursor.close()
        conn.close()
        
        if request.method == "POST":
            # รับค่าจากฟอร์มค้นหา
            start_date = request.form.get('start_date', '')
            end_date = request.form.get('end_date', '')
            booking_no = request.form.get('booking_no', '')
            name_surname = request.form.get('name_surname', '')
            
            # เชื่อมต่อกับฐานข้อมูล
            conn = get_db_connection()
            cursor = conn.cursor(cursor_factory=RealDictCursor)
            
            query = """
            SELECT 
                *
            FROM transfer_rental
            WHERE 1=1
            """
            
            params = []
            
            # เพิ่มเงื่อนไขการค้นหา
            if start_date:
                query += " AND travel_date >= %s"
                params.append(start_date)
            
            if end_date:
                query += " AND travel_date <= %s"
                params.append(end_date)
            
            if booking_no:
                query += " AND booking_no = %s"
                params.append(booking_no)
                
            if name_surname:
                # แยกชื่อและนามสกุล
                name_parts = name_surname.strip().split(' ', 1)
                if len(name_parts) == 2:
                    first_name, last_name = name_parts
                    query += " AND customer_name = %s AND customer_surname = %s"
                    params.extend([first_name, last_name])
                else:
                    # ถ้าไม่สามารถแยกได้ ให้ค้นหาจากชื่อเต็ม
                    query += " AND CONCAT(customer_name, ' ', customer_surname) = %s"
                    params.append(name_surname)
            
            query += " ORDER BY travel_date DESC, booking_no DESC"
            
            cursor.execute(query, params)
            bookings = cursor.fetchall()
            
            # แปลงรูปแบบวันที่และเวลา (เหมือนเดิม)
            for booking in bookings:
                if booking['travel_date']:
                    booking['travel_date'] = booking['travel_date'].strftime('%d/%m/%Y')
                if booking['booking_date']:
                    booking['booking_date'] = booking['booking_date'].strftime('%d/%m/%Y')
                
                if booking['pickup_time']:
                    if hasattr(booking['pickup_time'], 'strftime'):
                        booking['pickup_time'] = booking['pickup_time'].strftime('%H:%M')
                    else:
                        total_seconds = int(booking['pickup_time'].total_seconds())
                        hours = total_seconds // 3600
                        minutes = (total_seconds % 3600) // 60
                        booking['pickup_time'] = f"{hours:02d}:{minutes:02d}"
            
            cursor.close()
            conn.close()
            
            # ดึงข้อมูลเพิ่มเติมสำหรับ dropdown ใน edit modal
            conn = get_db_connection()
            cursor = conn.cursor(cursor_factory=RealDictCursor)
            
            departure_query = """
            SELECT DISTINCT place_to as destination, passengers, received, paid 
            FROM transfer_tabel 
            WHERE place_from = 'Lamai Bayview Boutique Resort'
            ORDER BY place_to
            """
            cursor.execute(departure_query)
            departure_destinations = cursor.fetchall()
            
            arrival_query = """
            SELECT DISTINCT place_from as origin, passengers, received, paid 
            FROM transfer_tabel 
            WHERE place_to = 'Lamai Bayview Boutique Resort'
            ORDER BY place_from
            """
            cursor.execute(arrival_query)
            arrival_origins = cursor.fetchall()
            
            cursor.close()
            conn.close()
            
            # ✅ เพิ่ม booking_list ในการ return
            return render_template('P2_home_transfer_form.html', 
                                 transfers=bookings,
                                 booking_list=booking_list,  # ✅ เพิ่มบรรทัดนี้
                                 departure_destinations=departure_destinations,
                                 arrival_origins=arrival_origins,
                                 search_start_date=start_date, 
                                 search_end_date=end_date, 
                                 search_booking_no=booking_no,
                                 search_name_surname=name_surname)
        else:
            # ✅ GET request - ใช้ booking_list ที่ดึงมาแล้ว
            return redirect(url_for('home_transfer_form'))
    
    except Exception as e:
        return f"Error: {str(e)}"

# Add route for canceling transfer bookings
@app.route("/cancel_transfer_bookings", methods=["POST"])
@login_required
def cancel_transfer_bookings():
    try:
        # รับค่า booking_no ที่ถูกเลือกจากฟอร์ม
        selected_bookings = request.form.getlist('selected_bookings')
        
        # รับชื่อผู้ cancel
        cancelled_by = request.form.get('cancelled_by', '').strip()
        
        if not selected_bookings:
            return jsonify({"success": False, "message": "No bookings selected"})
        
        if not cancelled_by:
            return jsonify({"success": False, "message": "Cancellation name is required"})
        
        # เชื่อมต่อกับฐานข้อมูล
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # อัพเดต payment_status แทนการลบ
        placeholders = ', '.join(['%s'] * len(selected_bookings))
        query = f"UPDATE transfer_rental SET payment_status = %s WHERE booking_no IN ({placeholders})"
        
        # สร้าง status ใหม่
        cancel_status = f"Cancelled by {cancelled_by}"
        params = [cancel_status] + selected_bookings
        
        cursor.execute(query, params)
        conn.commit()
        
        updated_count = cursor.rowcount
        
        cursor.close()
        conn.close()
        
        if updated_count > 0:
            return jsonify({"success": True, "message": f"Successfully cancelled {updated_count} booking(s) by {cancelled_by}"})
        else:
            return jsonify({"success": False, "message": "No bookings were cancelled"})
        
    except Exception as e:
        return jsonify({"success": False, "message": f"Error: {str(e)}"})

# Add route for getting transfer booking details for editing
@app.route("/get_transfer_booking_details", methods=["POST"])
@login_required
def get_transfer_booking_details():
    try:
        booking_no = request.form.get('booking_no')
        
        if not booking_no:
            return jsonify({"success": False, "message": "No booking selected"})
        
        conn = get_db_connection()
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        
        # เปลี่ยน JOIN condition
        query = """
        SELECT 
            tr.*,
            COALESCE(tt.paid, 0) as paid_from_table
        FROM transfer_rental tr
        LEFT JOIN transfer_tabel tt ON tr.place_from = tt.place_from 
                                    AND tr.place_to = tt.place_to
        WHERE tr.booking_no = %s
        """
        cursor.execute(query, (booking_no,))
        booking = cursor.fetchone()
        
        cursor.close()
        conn.close()
        
        if not booking:
            return jsonify({"success": False, "message": "Booking not found"})
        
        # Use paid amount from transfer_tabel if available, otherwise use 0
        booking['paid'] = booking['paid_from_table'] if booking['paid_from_table'] is not None else 0
        
        # Format dates and times
        if booking['travel_date']:
            booking['travel_date'] = booking['travel_date'].isoformat()
        if booking['booking_date']:
            booking['booking_date'] = booking['booking_date'].isoformat()
        
        # Format pickup time
        if booking['pickup_time']:
            if hasattr(booking['pickup_time'], 'strftime'):
                booking['pickup_time'] = booking['pickup_time'].strftime('%H:%M')
            else:
                # If it's a timedelta
                total_seconds = int(booking['pickup_time'].total_seconds())
                hours = total_seconds // 3600
                minutes = (total_seconds % 3600) // 60
                booking['pickup_time'] = f"{hours:02d}:{minutes:02d}"
        
        return jsonify({"success": True, "booking": booking})
        
    except Exception as e:
        return jsonify({"success": False, "message": f"Error: {str(e)}"})

# Add route for updating transfer bookings
@app.route("/update_transfer_booking", methods=["POST"])
@login_required
def update_transfer_booking():
    try:
        # Get form data
        booking_no = request.form.get('booking_no')
        customer_name = request.form.get('name', '')
        customer_surname = request.form.get('surname', '')
        place_from = request.form.get('place_from', '')
        place_to = request.form.get('place_to', '')
        transfer_type = request.form.get('edit_transferType', '')  # เปลี่ยนจาก departure/arrivals
        detail = request.form.get('detail', '')
        pickup_time = request.form.get('time', '')
        travel_date = request.form.get('travel_date', '')
        quantity = request.form.get('persons', '0')
        received = request.form.get('received', '0')
        payment_status = request.form.get('status', '')
        staff_name = request.form.get('staffName', '')
        driver_name = request.form.get('driverName', '')
        price = request.form.get('price', '0')
        payment_method = request.form.get('paymentmethod', '')
        remark = request.form.get('remark', '')
        discount = request.form.get('discount', '')
        
        if not booking_no:
            return jsonify({"success": False, "message": "Booking number is required"})
        
        # Connect to database
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # อัพเดท query ใหม่ - เปลี่ยนจาก departure, arrivals เป็น transfer_type
        query = """
        UPDATE transfer_rental SET
            travel_date = %s,
            pickup_time = %s,
            customer_name = %s,
            customer_surname = %s,
            place_from = %s,
            place_to = %s,
            transfer_type = %s,
            detail = %s,
            quantity = %s,
            received = %s,
            payment_status = %s,
            staff_name = %s,
            driver_name = %s,
            price = %s,
            payment_method = %s,
            remark = %s,
            discount = %s
        WHERE booking_no = %s
        """
        
        values = (
            travel_date, pickup_time, customer_name, customer_surname,
            place_from, place_to, transfer_type, detail, 
            quantity, received, payment_status, staff_name, 
            driver_name, price, payment_method, remark, discount, booking_no
        )
        
        cursor.execute(query, values)
        conn.commit()
        
        updated = cursor.rowcount > 0
        
        cursor.close()
        conn.close()
        
        if updated:
            return jsonify({"success": True, "message": "Transfer booking successfully updated"})
        else:
            return jsonify({"success": False, "message": "No changes were made or booking not found"})
        
    except Exception as e:
        return jsonify({"success": False, "message": f"Error: {str(e)}"})
    
    
# ---- EXCEL FORM GENERATION FUNCTIONALITY ----

@app.route("/generate_excel_form_transfer", methods=["POST"])
def generate_excel_form_transfer():
    try:
        booking_no = request.form.get('booking_no')
        
        if not booking_no:
            return jsonify({"success": False, "message": "Booking not found"}), 400
        
        print(f"Processing Excel generation for transfer booking: {booking_no}")
        
        # เชื่อมต่อฐานข้อมูล
        conn = get_db_connection()
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        
        # ดึงข้อมูลการจอง
        query = """
        SELECT 
            booking_no, booking_date, travel_date, pickup_time, 
            customer_name, customer_surname,
            CONCAT(customer_name, ' ', customer_surname) AS full_name,
            place_from, place_to, transfer_type, detail,
            quantity, received, payment_status, staff_name,
            driver_name, price, payment_method, remark, discount
        FROM transfer_rental 
        WHERE booking_no = %s
        """
        cursor.execute(query, (booking_no,))
        booking = cursor.fetchone()
        
        if not booking:
            cursor.close()
            conn.close()
            return jsonify({"success": False, "message": f"Booking not found: {booking_no}"}), 404
        
        print(f"Transfer booking data retrieved: {booking}")
        
        # สร้างไฟล์ Excel
        excel_file = create_excel_form_transfer(booking)
        print(f"Transfer Excel file created at: {excel_file}")
        
        # ⚠️ แปลงเป็น PDF เหมือนกับ Tour/Motorbike
        pdf_file = convert_excel_to_pdf(excel_file)
        print(f"Transfer PDF file: {pdf_file}")
        
        cursor.close()
        conn.close()
        
        # ตรวจสอบว่าได้ PDF หรือ Excel
        if pdf_file.endswith('.pdf'):
            return send_file(
                pdf_file,
                as_attachment=False,  # เปิดในเบราว์เซอร์
                download_name=f"{datetime.now().strftime('%Y%m%d')}_{booking_no}_Transfer_Booking.pdf",
                mimetype='application/pdf'
            )
        else:
            # ถ้าแปลง PDF ไม่ได้ ส่ง Excel แทน
            return send_file(
                excel_file,
                as_attachment=True,
                download_name=f"{datetime.now().strftime('%Y%m%d')}_{booking_no}_Transfer_Booking.xlsx",
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        
    except Exception as e:
        print(f"General error in transfer form generation: {str(e)}")
        return jsonify({"success": False, "message": f"Error: {str(e)}"}), 500
    
# ⚠️ ฟังก์ชันทำความสะอาดไฟล์ชั่วคราว (เพิ่มเติม)
def cleanup_transfer_temp_files():
    """ลบไฟล์ชั่วคราวของ Transfer เก่า"""
    temp_dir = tempfile.gettempdir()
    current_time = datetime.now()
    
    for filename in os.listdir(temp_dir):
        if filename.endswith(('.xlsx', '.pdf')) and 'transfer' in filename.lower():
            file_path = os.path.join(temp_dir, filename)
            try:
                # ลบไฟล์ที่เก่ากว่า 1 ชั่วโมง
                file_time = datetime.fromtimestamp(os.path.getctime(file_path))
                if (current_time - file_time).seconds > 3600:
                    os.remove(file_path)
                    print(f"Cleaned up transfer temp file: {filename}")
            except Exception as e:
                print(f"Error cleaning transfer temp file {filename}: {e}")

def create_excel_form_transfer(booking):
    """Create Excel form from transfer booking data with images"""
    # Define template path in static folder
    script_dir = os.path.dirname(os.path.abspath(__file__))
    template_path = os.path.join(script_dir, 'static', 'templates', 'transfer_template.xlsx')
    
    # ตรวจสอบว่าไฟล์ template มีอยู่จริง
    if not os.path.exists(template_path):
        print(f"Transfer template not found at: {template_path}")
        
        templates_dir = os.path.join(script_dir, 'static', 'templates')
        if not os.path.exists(templates_dir):
            os.makedirs(templates_dir)
            print(f"Created templates directory: {templates_dir}")
        
        raise FileNotFoundError(f"Could not find the Transfer Excel template file at: {template_path}")
    
    # โหลด Excel template
    workbook = openpyxl.load_workbook(template_path)
    sheet = workbook.active
    
    # Cell mappings ตามที่กำหนด
    cell_mappings = {
        "booking_date": ["J3", "J30"],
        "booking_no": ["J4", "J31"],
        "customer_name": ["D11", "D38", "G25", "G52"],
        "place_from": ["D13", "D40"],
        "place_to": ["D15", "D42"],
        "travel_date": ["H13", "H40"],
        "pickup_time": ["H15", "H42"],
        "quantity": ["E17", "E44"],
        "received": ["H17", "H44"],
        "detail": ["D19", "D46"],
        "payment_status": ["D22", "D49"],
        "staff_name": ["B25", "B52"],
        "remark": ["D21", "D48"],
        "payment_method": ["H22", "H49"]
    }
    
    # ใส่ข้อมูลลงในเซลล์ตาม mappings
    for field, cells in cell_mappings.items():
        value = None
        
        # กรณีพิเศษสำหรับชื่อลูกค้าเต็ม
        if field == "customer_name":
            value = booking["full_name"]
        # ✅ เพิ่ม capitalize สำหรับ payment_status
        elif field == "payment_status":
            if booking.get(field):
                original_value = booking[field]
                value = str(booking[field]).title()  # ใช้ title() แทน capitalize()
                print(f"DEBUG transfer payment_status: '{original_value}' → '{value}'")
            else:
                value = ""
                print("DEBUG transfer payment_status: No value found")
        elif field in booking and booking[field] is not None:
            value = booking[field]
        
        # จัดรูปแบบวันที่และเวลา
        if field == "booking_date" or field == "travel_date":
            if value:
                if isinstance(value, (date, datetime)):
                    value = value.strftime("%d/%m/%Y")
        elif field == "pickup_time" and value:
            if isinstance(value, (time, datetime)):
                value = value.strftime("%H:%M")
            elif hasattr(value, 'total_seconds'):
                # กรณีเป็น timedelta
                total_seconds = int(value.total_seconds())
                hours = total_seconds // 3600
                minutes = (total_seconds % 3600) // 60
                value = f"{hours:02d}:{minutes:02d}"
            
        # ใส่ข้อมูลลงในทุกเซลล์ที่เกี่ยวข้อง
        if value is not None:
            for cell in cells:
                sheet[cell] = value
    
    # *** เพิ่มส่วนรูปภาพเหมือนกับ Tour/Motorbike ***
    logo_path = os.path.join(script_dir, 'static', 'images', 'logoexcel.png')
    
    if os.path.exists(logo_path):
        try:
            # เพิ่มโลโก้ที่ B1 (หน้าแรก)
            img1 = Image(logo_path)
            img1.width = 100   # ปรับขนาดตามต้องการ
            img1.height = 80
            img1.anchor = 'B1'
            sheet.add_image(img1)
            
            # เพิ่มโลโก้ที่ B28 (หน้าที่สอง)
            img2 = Image(logo_path)
            img2.width = 100
            img2.height = 80
            img2.anchor = 'B28'
            sheet.add_image(img2)
            
            print(f"Transfer logos added successfully")
            
        except Exception as e:
            print(f"Error adding transfer logos: {str(e)}")
    else:
        print(f"Transfer logo file not found at: {logo_path}")
    
    # เพิ่มรูปภาพอื่นๆ สำหรับ Transfer
    additional_images = [
        {'path': os.path.join(script_dir, 'static', 'images', 'signature.png'), 'anchor': 'H25'},
        {'path': os.path.join(script_dir, 'static', 'images', 'signature.png'), 'anchor': 'H52'},  # หน้าที่สอง
    ]
    
    for img_info in additional_images:
        if os.path.exists(img_info['path']):
            try:
                img = Image(img_info['path'])
                img.width = 80
                img.height = 40
                img.anchor = img_info['anchor']
                sheet.add_image(img)
                print(f"Transfer additional image added")
            except Exception as e:
                print(f"Error adding transfer image: {str(e)}")
    
    # สร้างไฟล์ชั่วคราว
    temp_dir = tempfile.gettempdir()
    temp_file = os.path.join(temp_dir, f"{uuid.uuid4()}.xlsx")
    
    # บันทึก workbook ไปยังไฟล์ชั่วคราว
    workbook.save(temp_file)
    
    return temp_file

@app.route("/export_transfers", methods=["POST"])
@login_required
def export_transfers():
    try:
        # Get filter parameters
        filter_type = request.form.get('filter_type', '')
        transfer_type = request.form.get('transfer_type', 'all')
        payment_status = request.form.get('payment_status', 'all')
        
        # Base query - Updated to match new schema with JOIN to get paid amount
        query = """
        SELECT 
            tr.booking_date,
            tr.booking_no, 
            tr.travel_date,
            tr.customer_name, 
            tr.customer_surname, 
            tr.place_from, 
            tr.place_to, 
            tr.detail, 
            tr.staff_name,
            tr.driver_name,
            tr.quantity, 
            tr.received,
            tr.payment_status,
            tr.transfer_type,  -- เปลี่ยนจาก departure, arrivals
            tr.price,
            tr.payment_method,
            tr.remark,
            tr.discount,
            COALESCE(tt.paid, 0) as paid
        FROM transfer_rental tr
        LEFT JOIN transfer_tabel tt ON tr.place_from = tt.place_from 
                                    AND tr.place_to = tt.place_to
        WHERE 1=1
        """
        
        params = []
        
        # Apply date filters based on filter type
        if filter_type == 'date':
            start_date = request.form.get('start_date', '')
            end_date = request.form.get('end_date', '')
            
            if start_date:
                query += " AND tr.travel_date >= %s"
                params.append(start_date)
            
            if end_date:
                query += " AND tr.travel_date <= %s"
                params.append(end_date)
                
        elif filter_type == 'month':
            start_month = request.form.get('start_month', '')
            end_month = request.form.get('end_month', '')
            
            if start_month:
                start_date = f"{start_month}-01"
                query += " AND tr.travel_date >= %s"
                params.append(start_date)
            
            if end_month:
                year, month = map(int, end_month.split('-'))
                if month == 12:
                    next_year = year + 1
                    next_month = 1
                else:
                    next_year = year
                    next_month = month + 1
                
                end_date = f"{next_year}-{next_month:02d}-01"
                query += " AND tr.travel_date < %s"
                params.append(end_date)
                
        elif filter_type == 'year':
            start_year = request.form.get('start_year', '')
            end_year = request.form.get('end_year', '')
            
            if start_year:
                start_date = f"{start_year}-01-01"
                query += " AND tr.travel_date >= %s"
                params.append(start_date)
            
            if end_year:
                end_date = f"{int(end_year) + 1}-01-01"
                query += " AND tr.travel_date < %s"
                params.append(end_date)
        
        # Apply transfer type filter
        if transfer_type == 'departure':
            query += " AND tr.transfer_type = 'departure'"
        elif transfer_type == 'arrivals':
            query += " AND tr.transfer_type = 'arrivals'"
        # 'all' = no additional filter
        
        # Apply payment status filter
        if payment_status == 'paid':
            query += " AND tr.payment_status = 'paid'"
        elif payment_status == 'unpaid':
            query += " AND tr.payment_status = 'unpaid'"
        elif payment_status == 'partial':
            query += " AND tr.payment_status = 'partial'"
        # 'all' = no additional filter
        
        # Order by date and booking number
        query += " ORDER BY tr.booking_date DESC, tr.booking_no DESC"
        
        # Connect to database
        conn = get_db_connection()
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        
        # Execute query
        cursor.execute(query, params)
        transfers = cursor.fetchall()
        
        cursor.close()
        conn.close()
        
        # Create Excel file
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        
        # Set sheet title based on transfer type
        if transfer_type == 'departure':
            sheet.title = "Departure Transfers"
        elif transfer_type == 'arrivals':
            sheet.title = "Arrival Transfers"
        else:
            sheet.title = "Transfer Bookings"
        
        # Add headers - Updated to match your requested format
        headers = [
            "Booking date", "Booking No.", "Name&Surname", 
            "Departure/Arrivals", "From", "To", "Detail", 
            "Person", "Price", "Discount", "Selling Price"
            ]

        if session.get('role') == 'admin':    
            headers.extend(["Cost", "Amount"])

        headers.extend([    
            "Payment Status", "Staff Name", "Driver name", "Payment Method", "Remark"
        ])
        
        # Style headers
        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num)
            cell.value = header
            cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
            cell.fill = openpyxl.styles.PatternFill(
                start_color="366092", end_color="366092", fill_type="solid"
            )
            cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
        
        # Add data rows
        for row_num, transfer in enumerate(transfers, 2):
            # Format booking date
            booking_date = transfer['booking_date']
            if isinstance(booking_date, (date, datetime)):
                booking_date = booking_date.strftime("%d/%m/%Y")
            
            # ✅ ดึง price มาตรงๆ จาก database
            price = float(transfer['price']) if transfer['price'] else 0.0
            
            # Calculate amount (difference between received and paid)
            paid_amount = float(transfer['paid']) if transfer['paid'] else 0.0
            received_amount = float(transfer['received']) if transfer['received'] else 0.0
            if transfer['payment_status'].lower().startswith('cancelled'):
                amount = 0  # cancel = 0
            else:
                amount = received_amount - paid_amount  # paid หรือ unpaid = received - paid
            
            # Full name
            full_name = f"{transfer['customer_name']} {transfer['customer_surname']}"
            
            # Transfer type display
            transfer_type_text = "Unknown"
            if transfer['transfer_type'] == 'departure':
                transfer_type_text = "Departure"
            elif transfer['transfer_type'] == 'arrivals':
                transfer_type_text = "Arrivals"
            
            
            # Add row data
            row_data = [
                booking_date,                    # Booking date
                transfer['booking_no'],          # Booking No.
                full_name,                       # Name&Surname
                transfer_type_text,              # Departure/Arrivals
                transfer['place_from'],          # From
                transfer['place_to'],            # To
                transfer['detail'] or '',        # Detail
                transfer['quantity'], 
                price,
                transfer['discount'],
                received_amount  
            ]
            
            if session.get('role') == 'admin':
                row_data.extend([paid_amount, round(amount, 2)])

            row_data.extend([
                transfer['payment_status'],
                transfer['staff_name'], 
                transfer['driver_name'] or '',   # Driver name
                transfer['payment_method'],      # Payment Method
                transfer['remark'],              # Remark
            ])
            
            for col_num, cell_value in enumerate(row_data, 1):
                cell = sheet.cell(row=row_num, column=col_num)
                cell.value = cell_value
                
                # Add alternating row colors for better readability
                if row_num % 2 == 0:
                    cell.fill = openpyxl.styles.PatternFill(
                        start_color="F8F9FA", end_color="F8F9FA", fill_type="solid"
                    )
        
        # Auto-size columns
        for column in sheet.columns:
            max_length = 0
            column_letter = openpyxl.utils.get_column_letter(column[0].column)
            for cell in column:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            sheet.column_dimensions[column_letter].width = adjusted_width

        # Add TOTAL row at the end of data
        if transfers and session.get('role') == 'admin':  # ✅ เพิ่มการตรวจสอบว่ามีข้อมูลหรือไม่
            total_row = len(transfers) + 2
            sheet.cell(row=total_row, column=10).value = "TOTAL"  # ✅ เปลี่ยนจาก column 11 เป็น 10 (Person column)
            sheet.cell(row=total_row, column=10).font = openpyxl.styles.Font(bold=True)

            # ✅ SUM formulas for columns 11, 12, 13, 14 (Price, Received, Paid, Amount)
            start_data_row = 2
            end_data_row = total_row - 1
            
            # Price total
            sheet.cell(row=total_row, column=11).value = f"=SUM(K{start_data_row}:K{end_data_row})"
            sheet.cell(row=total_row, column=11).font = openpyxl.styles.Font(bold=True)
            
            # Received total
            sheet.cell(row=total_row, column=12).value = f"=SUM(L{start_data_row}:L{end_data_row})"
            sheet.cell(row=total_row, column=12).font = openpyxl.styles.Font(bold=True)
            
            # Paid total
            sheet.cell(row=total_row, column=13).value = f"=SUM(M{start_data_row}:M{end_data_row})"
            sheet.cell(row=total_row, column=13).font = openpyxl.styles.Font(bold=True)
            
            
            # ✅ Optional styling for TOTAL row
            for col in range(10, 14):  # columns 10-14
                cell = sheet.cell(row=total_row, column=col)
                cell.fill = openpyxl.styles.PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
                cell.font = openpyxl.styles.Font(bold=True)

        # Create temporary file
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        workbook.save(temp_file.name)
        temp_file.close()
        
        # Send the file
        return send_file(
            temp_file.name,
            as_attachment=True,
            download_name=f"{datetime.now().strftime('%Y%m%d')}_Transfers_Export.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        
    except Exception as e:
        print(f"Export error: {str(e)}")
        return jsonify({"success": False, "message": f"Error: {str(e)}"}), 500
    
# ---------------------------------- For update Tabel ----------------------------------------------
# เพิ่ม routes เหล่านี้ใน Flask app ของคุณ

@app.route("/data_management")
@login_required
def data_management():
        # ตรวจสอบว่าเป็น admin หรือไม่
    if session.get('role') != 'admin':
        flash('Access Denied: Admin only feature', 'error')
        return redirect(url_for('home'))
        
    return render_template("P3_data_management.html")

# API สำหรับ Tour Data
@app.route("/api/tour_data", methods=["GET"])
@login_required
def get_tour_data():

    if session.get('role') != 'admin':
        return jsonify({"success": False, "message": "Access Denied: Admin only"})

    try:
        conn = get_db_connection()
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        
        query = "SELECT * FROM tour_tabel ORDER BY id"
        cursor.execute(query)
        data = cursor.fetchall()
        
        cursor.close()
        conn.close()
        
        return jsonify({"success": True, "data": data})
    except Exception as e:
        return jsonify({"success": False, "message": f"Error: {str(e)}"})

@app.route("/api/update_tour_data", methods=["POST"])
@login_required
def update_tour_data():

    
    if session.get('role') != 'admin':
        return jsonify({"success": False, "message": "Access Denied: Admin only"})

    try:
        data = request.json.get('data', [])
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # ลบข้อมูลทั้งหมดในตาราง
        cursor.execute("DELETE FROM tour_tabel")
        
        # เพิ่มข้อมูลใหม่
        for item in data:
            query = """
            INSERT INTO tour_tabel (company_name, detail, received, paid) 
            VALUES (%s, %s, %s, %s)
            """
            cursor.execute(query, (
                item['company_name'], 
                item['detail'], 
                item['received'], 
                item['paid']
            ))
        
        conn.commit()
        cursor.close()
        conn.close()
        
        return jsonify({"success": True, "message": "Tour data updated successfully"})
    except Exception as e:
        return jsonify({"success": False, "message": f"Error: {str(e)}"})

# API สำหรับ Motorbike Data
@app.route("/get_motorbike_data", methods=["GET"])
@login_required
def get_motorbike_data():
    """API endpoint สำหรับดึงข้อมูล motorbike companies และ details"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        
        # ดึงข้อมูลทั้งหมดจาก motorbike_tabel
        query = "SELECT company_name, detail, received, paid FROM motorbike_tabel ORDER BY company_name, detail"
        cursor.execute(query)
        data = cursor.fetchall()
        
        cursor.close()
        conn.close()
        
        # จัดระเบียบข้อมูลตาม company
        organized_data = {}
        for item in data:
            company = item['company_name']
            if company not in organized_data:
                organized_data[company] = {}
            
            organized_data[company][item['detail']] = {
                'price': float(item['received']) if item['received'] else 0,
                'paid': float(item['paid']) if item['paid'] else 0
            }
        
        return jsonify({
            "success": True, 
            "data": organized_data,
            "companies": list(organized_data.keys())
        })
        
    except Exception as e:
        return jsonify({
            "success": False, 
            "message": f"Error: {str(e)}",
            "data": {},
            "companies": []
        })

@app.route("/api/update_motorbike_data", methods=["POST"])
@login_required
def update_motorbike_data():

    
    if session.get('role') != 'admin':
        return jsonify({"success": False, "message": "Access Denied: Admin only"})

    try:
        data = request.json.get('data', [])
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # ลบข้อมูลทั้งหมดในตาราง
        cursor.execute("DELETE FROM motorbike_tabel")
        
        # เพิ่มข้อมูลใหม่
        for item in data:
            query = """
            INSERT INTO motorbike_tabel (company_name, detail, received, paid) 
            VALUES (%s, %s, %s, %s)
            """
            cursor.execute(query, (
                item['company_name'], 
                item['detail'], 
                item['received'], 
                item['paid']
            ))
        
        conn.commit()
        cursor.close()
        conn.close()
        
        return jsonify({"success": True, "message": "Motorbike data updated successfully"})
    except Exception as e:
        return jsonify({"success": False, "message": f"Error: {str(e)}"})

# API สำหรับ Transfer Data
@app.route("/api/transfer_data", methods=["GET"])
@login_required
def get_transfer_data():

    
    if session.get('role') != 'admin':
        return jsonify({"success": False, "message": "Access Denied: Admin only"})

    try:
        conn = get_db_connection()
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        
        query = "SELECT * FROM transfer_tabel ORDER BY id"
        cursor.execute(query)
        data = cursor.fetchall()
        
        cursor.close()
        conn.close()
        
        return jsonify({"success": True, "data": data})
    except Exception as e:
        return jsonify({"success": False, "message": f"Error: {str(e)}"})

@app.route("/api/update_transfer_data", methods=["POST"])
@login_required
def update_transfer_data():

    
    if session.get('role') != 'admin':
        return jsonify({"success": False, "message": "Access Denied: Admin only"})

    try:
        data = request.json.get('data', [])
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # ลบข้อมูลทั้งหมดในตาราง
        cursor.execute("DELETE FROM transfer_tabel")
        
        # เพิ่มข้อมูลใหม่
        for item in data:
            query = """
            INSERT INTO transfer_tabel (place_from, place_to, passengers, received, paid) 
            VALUES (%s, %s, %s, %s, %s)
            """
            cursor.execute(query, (
                item['place_from'], 
                item['place_to'], 
                item['passengers'], 
                item['received'], 
                item['paid']
            ))
        
        conn.commit()
        cursor.close()
        conn.close()
        
        return jsonify({"success": True, "message": "Transfer data updated successfully"})
    except Exception as e:
        return jsonify({"success": False, "message": f"Error: {str(e)}"})
    
@app.route("/api/motorbike_data", methods=["GET"])
@login_required  
def get_motorbike_data_api():
    if session.get('role') != 'admin':
        return jsonify({"success": False, "message": "Access Denied: Admin only"})

    try:
        conn = get_db_connection()
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        
        query = "SELECT * FROM motorbike_tabel ORDER BY id"
        cursor.execute(query)
        data = cursor.fetchall()
        
        cursor.close()
        conn.close()
        
        return jsonify({"success": True, "data": data})
    except Exception as e:
        return jsonify({"success": False, "message": f"Error: {str(e)}"})
    
# API สำหรับ Login Data
@app.route("/api/login_data", methods=["GET"])
@login_required
def get_login_data():

    
    if session.get('role') != 'admin':
        return jsonify({"success": False, "message": "Access Denied: Admin only"})

    try:
        conn = get_db_connection()
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        
        query = "SELECT * FROM login ORDER BY id"
        cursor.execute(query)
        data = cursor.fetchall()
        
        cursor.close()
        conn.close()
        
        return jsonify({"success": True, "data": data})
    except Exception as e:
        return jsonify({"success": False, "message": f"Error: {str(e)}"})


# API Login

@app.route("/api/update_login_data", methods=["POST"])
@login_required
def update_login_data():

    
    if session.get('role') != 'admin':
        return jsonify({"success": False, "message": "Access Denied: Admin only"})

    try:
        data = request.json.get('data', [])
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # ลบข้อมูลทั้งหมดในตาราง
        cursor.execute("DELETE FROM login")
        
        # เพิ่มข้อมูลใหม่
        for item in data:
            query = """
            INSERT INTO login (username, pass, first_name, last_name, role) 
            VALUES (%s, %s, %s, %s, %s)
            """
            cursor.execute(query, (
                item['username'], 
                item['pass'], 
                item['first_name'], 
                item['last_name'], 
                item['role']
            ))
        
        conn.commit()
        cursor.close()
        conn.close()
        
        return jsonify({"success": True, "message": "Login data updated successfully"})
    except Exception as e:
        return jsonify({"success": False, "message": f"Error: {str(e)}"})
    
# API สำหรับ Room Data
@app.route("/api/room_data", methods=["GET"])
@login_required
def get_room_data():

    
    if session.get('role') != 'admin':
        return jsonify({"success": False, "message": "Access Denied: Admin only"})

    try:
        conn = get_db_connection()
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        
        query = "SELECT * FROM room ORDER BY room"
        cursor.execute(query)
        data = cursor.fetchall()
        
        cursor.close()
        conn.close()
        
        return jsonify({"success": True, "data": data})
    except Exception as e:
        return jsonify({"success": False, "message": f"Error: {str(e)}"})


# API Room

@app.route("/api/update_room_data", methods=["POST"])
@login_required
def update_room_data():

    
    if session.get('role') != 'admin':
        return jsonify({"success": False, "message": "Access Denied: Admin only"})

    try:
        data = request.json.get('data', [])
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # ลบข้อมูลทั้งหมดในตาราง
        cursor.execute("DELETE FROM room")
        
        # เพิ่มข้อมูลใหม่
        for item in data:
            query = """
            INSERT INTO room (room) 
            VALUES (%s)
            """
            cursor.execute(query, (str(item['room']),))
        
        conn.commit()
        cursor.close()
        conn.close()
        
        return jsonify({"success": True, "message": "Room data updated successfully"})
    except Exception as e:
        return jsonify({"success": False, "message": f"Error: {str(e)}"})

def test_libreoffice_installation():
    """ทดสอบว่า LibreOffice ติดตั้งและใช้งานได้หรือไม่"""
    
    try:
        result = subprocess.run(
            ['libreoffice', '--version'],
            capture_output=True,
            text=True,
            timeout=10
        )
        
        if result.returncode == 0:
            print(f"LibreOffice is available: {result.stdout.strip()}")
            return True
        else:
            print(f"LibreOffice not working: {result.stderr}")
            return False
            
    except FileNotFoundError:
        print("LibreOffice not installed")
        return False
    except Exception as e:
        print(f"Error testing LibreOffice: {str(e)}")
        return False

if __name__ == "__main__":   
    app.run()